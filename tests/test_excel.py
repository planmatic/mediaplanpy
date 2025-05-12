"""
Tests for the Excel integration module.
"""
import os
import tempfile
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
import pytest

from mediaplanpy.models import MediaPlan, Meta, Campaign, LineItem
from mediaplanpy.excel import (
    export_to_excel,
    import_from_excel,
    update_from_excel,
    validate_excel
)


@pytest.fixture
def sample_mediaplan_v1():
    """Create a sample media plan (v1.0.0) for testing."""
    return MediaPlan(
        meta=Meta(
            id="mediaplan_test",
            schema_version="v1.0.0",
            created_by="test@example.com",
            created_at=datetime(2025, 1, 1, 12, 0, 0),
            name="Test Media Plan",
            comments="Test media plan"
        ),
        campaign=Campaign(
            id="test_campaign",
            name="Test Campaign",
            objective="awareness",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
            budget_total=Decimal("100000"),
            audience_age_start=18,
            audience_age_end=34,
            audience_gender="Any",  # Set a valid value
            location_type="Country",
            locations=["United States"]
        ),
        lineitems=[
            LineItem(
                id="test_lineitem",
                name="Social Line Item",
                start_date=date(2025, 1, 1),
                end_date=date(2025, 6, 30),
                cost_total=Decimal("50000"),
                channel="social",
                vehicle="Facebook",
                partner="Meta",
                kpi="CPM"
            )
        ]
    )


class TestExcelExport:
    """Tests for Excel export functionality."""

    def test_export_to_excel(self, sample_mediaplan_v1):
        """Test exporting a media plan to Excel."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Export to Excel
            exported_path = sample_mediaplan_v1.export_to_excel_path(tmp_path)

            # Check file exists
            assert os.path.exists(exported_path)
            assert os.path.getsize(exported_path) > 0

        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_export_with_default_path(self, sample_mediaplan_v1):
        """Test exporting with a default path based on media plan ID."""
        try:
            # Export without specifying path
            exported_path = sample_mediaplan_v1.export_to_excel_path()

            # Check the path includes media plan ID rather than campaign ID
            assert sample_mediaplan_v1.meta.id in exported_path
            assert exported_path.endswith(".xlsx")

            # Check file exists
            assert os.path.exists(exported_path)
            assert os.path.getsize(exported_path) > 0

        finally:
            # Clean up
            try:
                os.unlink(exported_path)
            except:
                pass

    def test_export_without_documentation(self, sample_mediaplan_v1):
        """Test exporting without the documentation sheet."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Export to Excel without documentation
            exported_path = sample_mediaplan_v1.export_to_excel_path(tmp_path, include_documentation=False)

            # Check file exists
            assert os.path.exists(exported_path)

            # Import to verify documentation sheet is missing
            import openpyxl
            workbook = openpyxl.load_workbook(exported_path)
            assert "Documentation" not in workbook.sheetnames

        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


class TestExcelImport:
    """Tests for Excel import functionality."""

    def test_import_from_excel(self, sample_mediaplan_v1):
        """Test importing a media plan from Excel."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Export to Excel
            sample_mediaplan_v1.export_to_excel_path(tmp_path)

            # Import from Excel
            imported_plan = MediaPlan.import_from_excel_path(tmp_path)

            # Check imported data
            assert imported_plan.meta.schema_version == "v1.0.0"
            assert imported_plan.campaign.name == "Test Campaign"
            assert len(imported_plan.lineitems) == 1
            assert imported_plan.lineitems[0].name == "Social Line Item"
            assert imported_plan.lineitems[0].cost_total == 50000  # Note: precision may be lost in Excel

        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_update_from_excel(self, sample_mediaplan_v1):
        """Test updating a media plan from Excel."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Export to Excel
            sample_mediaplan_v1.export_to_excel_path(tmp_path)

            # Modify the Excel file
            import openpyxl
            workbook = openpyxl.load_workbook(tmp_path)

            # Update campaign name
            campaign_sheet = workbook["Campaign"]
            for row in range(1, 10):
                if campaign_sheet.cell(row=row, column=1).value == "Campaign Name:":
                    campaign_sheet.cell(row=row, column=2).value = "Updated Campaign Name"
                    break

            # Save changes
            workbook.save(tmp_path)

            # Create a copy of the original plan
            import copy
            updated_plan = copy.deepcopy(sample_mediaplan_v1)

            # Update from Excel
            updated_plan.update_from_excel_path(tmp_path)

            # Check updates
            assert updated_plan.campaign.name == "Updated Campaign Name"
            assert updated_plan.meta.id == sample_mediaplan_v1.meta.id  # Should preserve ID

        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


class TestExcelValidation:
    """Tests for Excel validation functionality."""

    def test_validate_valid_excel(self, sample_mediaplan_v1):
        """Test validating a valid Excel file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Export to Excel
            sample_mediaplan_v1.export_to_excel_path(tmp_path)

            # Validate Excel
            errors = MediaPlan.validate_excel(tmp_path)

            # Should have no errors
            assert not errors

        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_validate_invalid_excel(self):
        """Test validating an invalid Excel file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Create an invalid Excel file
            import openpyxl
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet["A1"] = "Invalid Excel File"
            workbook.save(tmp_path)

            # Validate Excel
            errors = MediaPlan.validate_excel(tmp_path)

            # Should have errors
            assert errors
            assert len(errors) > 0

        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)