"""
Excel exporter for mediaplanpy - Updated for v3.0 Schema Support Only.

This module provides functionality for exporting media plans to Excel format,
supporting only v3.0 schema with target audiences/locations arrays and all new fields.
"""

import os
import logging
from typing import Dict, Any, List, Optional, Union, Tuple, TYPE_CHECKING
from datetime import datetime
from decimal import Decimal

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell import Cell

from mediaplanpy.exceptions import StorageError
from mediaplanpy.workspace import WorkspaceManager

if TYPE_CHECKING:
    from mediaplanpy.models.mediaplan import MediaPlan
    from mediaplanpy.models.lineitem import LineItem
    from mediaplanpy.models.dictionary import Dictionary

logger = logging.getLogger("mediaplanpy.excel.exporter")


def export_to_excel(media_plan: "MediaPlan", path: Optional[str] = None,
                    template_path: Optional[str] = None,
                    include_documentation: bool = True,
                    workspace_manager: Optional[WorkspaceManager] = None,
                    **kwargs) -> str:
    """
    Export a media plan to Excel format using v3.0 schema.

    Args:
        media_plan: MediaPlan object to export (must be v3.0 schema)
        path: Optional path for the output file
        template_path: Optional path to an Excel template file
        include_documentation: Whether to include a documentation sheet
        workspace_manager: Optional WorkspaceManager for workspace storage
        **kwargs: Additional export options

    Returns:
        The path to the exported Excel file

    Raises:
        StorageError: If export fails or schema version is not v3.0
    """
    try:
        # Determine the path if not provided
        if not path:
            media_plan_id = media_plan.meta.id
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = f"{media_plan_id}_{timestamp}.xlsx"

        # Validate schema version - only v3.0 supported
        schema_version = media_plan.meta.schema_version
        if not _is_v3_schema_version(schema_version):
            raise StorageError(f"Excel export only supports v3.0 schema. Found: {schema_version}")

        # Create workbook
        if template_path and os.path.exists(template_path):
            workbook = openpyxl.load_workbook(template_path)
        else:
            workbook = _create_v3_workbook()

        # Populate the workbook with v3.0 data
        _populate_v3_workbook(workbook, media_plan, include_documentation)

        # Add validation and formatting
        _add_v3_validation_and_formatting(workbook)

        # Save the workbook
        if workspace_manager is not None:
            # Make sure workspace is loaded
            if not workspace_manager.is_loaded:
                workspace_manager.load()

            # Get storage backend
            storage_backend = workspace_manager.get_storage_backend()

            # Save to a temporary file first
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name
                workbook.save(tmp_path)

            # Read the content
            with open(tmp_path, 'rb') as f:
                content = f.read()

            # Write to storage backend
            storage_backend.write_file(path, content)

            # Clean up temp file
            os.unlink(tmp_path)

            logger.info(f"Media plan exported to Excel in workspace storage: {path}")
        else:
            # Save locally
            workbook.save(path)
            logger.info(f"Media plan exported to Excel at: {path}")

        return path

    except Exception as e:
        raise StorageError(f"Failed to export media plan to Excel: {e}")


def _is_v3_schema_version(version: str) -> bool:
    """
    Check if the schema version is v3.0.

    Args:
        version: The schema version to check

    Returns:
        True if the version is v3.0, False otherwise
    """
    if not version:
        return False

    # Normalize version format (handle both "3.0" and "v3.0")
    normalized = version.replace("v", "") if version.startswith("v") else version
    return normalized == "3.0"


def _create_v3_workbook() -> Workbook:
    """
    Create a default workbook with v3.0 schema sheets.

    Returns:
        A new Workbook with v3.0 schema sheets
    """
    workbook = Workbook()

    # Rename default sheet to "Metadata"
    metadata_sheet = workbook.active
    metadata_sheet.title = "Metadata"

    # Create other required sheets for v3.0
    campaign_sheet = workbook.create_sheet("Campaign")
    target_audiences_sheet = workbook.create_sheet("Target Audiences")  # NEW for v3.0
    target_locations_sheet = workbook.create_sheet("Target Locations")  # NEW for v3.0
    lineitems_sheet = workbook.create_sheet("Line Items")
    dictionary_sheet = workbook.create_sheet("Dictionary")
    documentation_sheet = workbook.create_sheet("Documentation")

    # Create v3.0 styles
    _create_v3_styles(workbook)

    return workbook


# ============================================================================
# Formula-Aware Column Building Helper Functions (NEW for v3.0 formulas)
# ============================================================================

def _get_formula_config(metric_name: str, dictionary: Dict[str, Any]) -> Optional[Dict[str, str]]:
    """
    Get formula configuration from Dictionary for a metric.

    Args:
        metric_name: Name of the metric (e.g., "metric_clicks")
        dictionary: Dictionary configuration from media plan

    Returns:
        Dictionary with formula_type and base_metric, or None if not configured:
        {
            "formula_type": "conversion_rate",
            "base_metric": "metric_impressions"
        }

    Examples:
        >>> _get_formula_config("metric_clicks", dictionary)
        {"formula_type": "conversion_rate", "base_metric": "metric_impressions"}

        >>> _get_formula_config("metric_impressions", {})
        None  # No dictionary config, will use defaults
    """
    if not dictionary:
        return None

    # Check standard_metrics first
    standard_metrics = getattr(dictionary,"standard_metrics", {})
    if metric_name in standard_metrics:
        config = standard_metrics[metric_name]
        # Extract formula_type and base_metric
        formula_type = getattr(config, 'formula_type', None) if hasattr(config, 'formula_type') else config.get('formula_type')
        base_metric = getattr(config, 'base_metric', None) if hasattr(config, 'base_metric') else config.get('base_metric')

        if formula_type and base_metric:
            return {
                "formula_type": formula_type,
                "base_metric": base_metric
            }

    # Check custom_metrics
    custom_metrics = getattr(dictionary,"custom_metrics", {})
    if metric_name in custom_metrics:
        config = custom_metrics[metric_name]
        # Extract formula_type and base_metric from dict
        formula_type = config.get('formula_type') if isinstance(config, dict) else None
        base_metric = config.get('base_metric') if isinstance(config, dict) else None

        if formula_type and base_metric:
            return {
                "formula_type": formula_type,
                "base_metric": base_metric
            }

    return None


def _determine_calculated_columns(
    metric_name: str,
    metric_header: str,
    formula_config: Optional[Dict[str, str]]
) -> List[Tuple[str, str, str]]:
    """
    Determine which calculated columns to create based on formula_type.

    Args:
        metric_name: Field name (e.g., "metric_clicks")
        metric_header: Display name (e.g., "Clicks")
        formula_config: Formula configuration from dictionary, or None for defaults

    Returns:
        List of (field_name, header_name, field_type) tuples for calculated and metric columns

    Examples:
        >>> # Default (no config): cost_per_unit
        >>> _determine_calculated_columns("metric_clicks", "Clicks", None)
        [("metric_clicks_cpu", "Cost per Click", "calculated"),
         ("metric_clicks", "Clicks", "formula")]

        >>> # Conversion rate
        >>> config = {"formula_type": "conversion_rate", "base_metric": "metric_impressions"}
        >>> _determine_calculated_columns("metric_clicks", "Clicks", config)
        [("metric_clicks_cvr", "Clicks Conversion Rate", "calculated"),
         ("metric_clicks", "Clicks", "formula")]

        >>> # Constant
        >>> config = {"formula_type": "constant", "base_metric": "cost_total"}
        >>> _determine_calculated_columns("metric_reach", "Reach", config)
        [("metric_reach_const", "Reach Constant", "calculated"),
         ("metric_reach", "Reach", "formula")]

        >>> # Power function
        >>> config = {"formula_type": "power_function", "base_metric": "metric_impressions"}
        >>> _determine_calculated_columns("metric_custom1", "Metric Custom 1", config)
        [("metric_custom1_coef", "Metric Custom 1 Coefficient", "calculated"),
         ("metric_custom1_param1", "Metric Custom 1 Parameter 1", "calculated"),
         ("metric_custom1", "Metric Custom 1", "formula")]
    """
    columns = []

    # Determine formula type (default to cost_per_unit)
    if formula_config is None:
        formula_type = "cost_per_unit"
    else:
        formula_type = formula_config.get("formula_type", "cost_per_unit")

    # Generate calculated column names and headers based on formula type
    if formula_type == "cost_per_unit":
        # Cost per unit: CPU column
        calc_field = f"{metric_name}_cpu"

        # Special handling for impressions (CPM)
        if metric_name == "metric_impressions":
            calc_header = "Cost per 1000 Impressions"
        else:
            # Convert "Clicks" -> "Click", but preserve special cases
            singular_name = metric_header
            if singular_name.endswith('s') and singular_name not in ['Views', 'Sales']:
                singular_name = singular_name.rstrip('s')
            calc_header = f"Cost per {singular_name}"

        columns.append((calc_field, calc_header, "calculated"))

    elif formula_type == "conversion_rate":
        # Conversion rate: CVR column
        calc_field = f"{metric_name}_cvr"
        calc_header = f"{metric_header} Conversion Rate"
        columns.append((calc_field, calc_header, "calculated"))

    elif formula_type == "constant":
        # Constant: No calculated column needed - coefficient goes directly in formula
        pass

    elif formula_type == "power_function":
        # Power function: coefficient and parameter1 columns
        coef_field = f"{metric_name}_coef"
        coef_header = f"{metric_header} Coefficient"
        columns.append((coef_field, coef_header, "calculated"))

        param1_field = f"{metric_name}_param1"
        param1_header = f"{metric_header} Parameter 1"
        columns.append((param1_field, param1_header, "calculated"))

    # Add the metric column itself (with formula)
    columns.append((metric_name, metric_header, "formula"))

    return columns


def _get_missing_base_metrics(
    present_metrics: List[Tuple[str, str]],
    dictionary: Dict[str, Any]
) -> List[Tuple[str, str]]:
    """
    Identify base metrics that are missing but needed for formulas.

    Per Design Decision 1: If a metric has a base_metric that doesn't exist in the
    line items, add that base_metric column with 0 values.

    Args:
        present_metrics: List of (field_name, header_name) tuples for metrics present in data
        dictionary: Dictionary configuration

    Returns:
        List of (field_name, header_name) tuples for missing base metrics to add

    Example:
        >>> present = [("metric_clicks", "Clicks"), ("metric_conversions", "Conversions")]
        >>> dictionary = {
        ...     "standard_metrics": {
        ...         "metric_clicks": {"formula_type": "conversion_rate", "base_metric": "metric_impressions"},
        ...         "metric_conversions": {"formula_type": "conversion_rate", "base_metric": "metric_clicks"}
        ...     }
        ... }
        >>> _get_missing_base_metrics(present, dictionary)
        [("metric_impressions", "Impressions")]  # Missing but needed for clicks formula
    """
    if not dictionary:
        return []

    missing = []
    present_field_names = {field_name for field_name, _ in present_metrics}

    # Check all present metrics to see if their base_metrics exist
    for metric_name, _ in present_metrics:
        formula_config = _get_formula_config(metric_name, dictionary)
        if formula_config:
            base_metric = formula_config.get("base_metric")
            if base_metric and base_metric not in present_field_names:
                # IMPORTANT: Skip cost fields - they are always added in base section
                # and should never have formulas. Only add missing metric fields.
                if base_metric.startswith("cost_"):
                    logger.debug(f"Skipping cost field {base_metric} - already in base fields")
                    continue

                # Base metric is missing - need to add it
                # Generate header name from field name
                if base_metric.startswith("metric_"):
                    # Convert metric_impressions -> Impressions
                    header = base_metric.replace("metric_", "").replace("_", " ").title()
                else:
                    # Convert cost_total -> Cost Total
                    header = base_metric.replace("_", " ").title()

                missing.append((base_metric, header))
                # Add to present set so we don't add duplicates
                present_field_names.add(base_metric)
                logger.info(f"Added missing base metric column: {base_metric}")

    return missing


def _populate_coefficient_column(
    metric_name: str,
    line_item: "LineItem",
    formula_config: Optional[Dict[str, str]]
) -> Optional[Decimal]:
    """
    Calculate coefficient value to populate in Excel calculated column.

    This function either retrieves the coefficient from lineitem.metric_formulas,
    or reverse-calculates it from the metric and base metric values.

    Args:
        metric_name: Name of the metric (e.g., "metric_clicks")
        line_item: LineItem object
        formula_config: Formula configuration with formula_type and base_metric

    Returns:
        Decimal coefficient value, or None if cannot be calculated

    Examples:
        >>> # Has formula with coefficient
        >>> line_item = {
        ...     "metric_clicks": 20000,
        ...     "metric_formulas": {
        ...         "metric_clicks": {"coefficient": Decimal("0.02")}
        ...     }
        ... }
        >>> _populate_coefficient_column("metric_clicks", line_item, config)
        Decimal("0.02")

        >>> # No formula, reverse-calculate from values (conversion_rate)
        >>> line_item = {
        ...     "metric_clicks": 20000,
        ...     "metric_impressions": 1000000
        ... }
        >>> config = {"formula_type": "conversion_rate", "base_metric": "metric_impressions"}
        >>> _populate_coefficient_column("metric_clicks", line_item, config)
        Decimal("0.02")  # 20000 / 1000000
    """
    # Step 1: Check if formula exists in lineitem with coefficient
    metric_formulas = getattr(line_item, "metric_formulas", {})
    if metric_formulas is None:
        metric_formulas = {}

    if metric_formulas and metric_name in metric_formulas:
        formula = metric_formulas[metric_name]
        # MetricFormula object (always an object in Pydantic models)
        coefficient = getattr(formula, "coefficient", None)
        formula_type_from_formula = getattr(formula, "formula_type", None)

        if coefficient is not None:
            # Convert to Decimal if needed
            if not isinstance(coefficient, Decimal):
                coefficient = Decimal(str(coefficient))

            # Special case: For impressions with cost_per_unit, multiply by 1000 to show CPM
            if metric_name == "metric_impressions" and formula_type_from_formula == "cost_per_unit":
                coefficient = coefficient * 1000

            return coefficient

    # Step 2: Reverse-calculate from metric and base metric values
    metric_value = getattr(line_item, metric_name, None)
    if metric_value is None or metric_value == 0:
        return Decimal("0")

    # Convert to Decimal
    if not isinstance(metric_value, Decimal):
        metric_value = Decimal(str(metric_value))

    # Get formula type and base metric (default to cost_per_unit/cost_total)
    if formula_config is None:
        formula_type = "cost_per_unit"
        base_metric_name = "cost_total"
    else:
        formula_type = formula_config.get("formula_type", "cost_per_unit")
        base_metric_name = formula_config.get("base_metric", "cost_total")

    # Get base metric value
    base_metric_value = getattr(line_item, base_metric_name, None)
    if base_metric_value is None or base_metric_value == 0:
        return Decimal("0")

    # Convert to Decimal
    if not isinstance(base_metric_value, Decimal):
        base_metric_value = Decimal(str(base_metric_value))

    # Reverse-calculate coefficient based on formula type
    try:
        if formula_type == "cost_per_unit":
            # coefficient = base / metric
            coefficient = base_metric_value / metric_value

            # Special case: For impressions, multiply by 1000 to show CPM
            # (Cost Per Mille = cost per 1000 impressions)
            if metric_name == "metric_impressions":
                coefficient = coefficient * 1000

        elif formula_type == "conversion_rate":
            # coefficient = metric / base
            coefficient = metric_value / base_metric_value

        elif formula_type == "constant":
            # coefficient = metric (constant value)
            coefficient = metric_value

        elif formula_type == "power_function":
            # coefficient = metric / (base ^ parameter1)
            # Get parameter1 from formula or default to 1.0
            parameter1 = Decimal("1.0")
            if metric_formulas and metric_name in metric_formulas:
                formula = metric_formulas[metric_name]
                # MetricFormula object (always an object in Pydantic models)
                param1 = getattr(formula, "parameter1", None)

                if param1 is not None:
                    parameter1 = Decimal(str(param1))

            # Calculate: coefficient = metric / (base ^ param1)
            if base_metric_value > 0:
                coefficient = metric_value / (base_metric_value ** parameter1)
            else:
                coefficient = Decimal("0")

        else:
            # Unknown formula type, default to 0
            coefficient = Decimal("0")

        return coefficient

    except (ZeroDivisionError, ValueError, ArithmeticError):
        # Handle any calculation errors
        return Decimal("0")


def _get_parameter1_value(
    metric_name: str,
    line_item: "LineItem"
) -> Decimal:
    """
    Get parameter1 value for power_function formulas.

    Args:
        metric_name: Name of the metric
        line_item: LineItem object

    Returns:
        Decimal parameter1 value, defaults to 1.0 if not found
    """
    metric_formulas = getattr(line_item, "metric_formulas", {})
    if metric_formulas is None:
        metric_formulas = {}

    if metric_formulas and metric_name in metric_formulas:
        formula = metric_formulas[metric_name]
        # MetricFormula object (always an object in Pydantic models)
        param1 = getattr(formula, "parameter1", None)

        if param1 is not None:
            if not isinstance(param1, Decimal):
                return Decimal(str(param1))
            return param1

    return Decimal("1.0")


def _generate_excel_formula(
    metric_name: str,
    formula_config: Optional[Dict[str, str]],
    column_refs: Dict[str, str],
    line_item: "LineItem"
) -> str:
    """
    Generate Excel formula string based on formula_type and base_metric.

    Args:
        metric_name: Name of the metric (e.g., "metric_clicks")
        formula_config: Formula configuration with formula_type and base_metric
        column_refs: Dictionary mapping field names to Excel column references
            Example: {"cost_total": "$C5", "metric_impressions": "$B5", "metric_clicks_cvr": "$D5"}
        line_item: Line item data dictionary (needed for constant formula coefficient)

    Returns:
        Excel formula string (e.g., "=IF($D5=0,0,$B5*$D5)")

    Examples:
        >>> # Cost per unit (default)
        >>> config = None
        >>> refs = {"cost_total": "$C5", "metric_clicks_cpu": "$D5"}
        >>> _generate_excel_formula("metric_clicks", config, refs)
        "=IF($D5=0,0,$C5/$D5)"

        >>> # Conversion rate
        >>> config = {"formula_type": "conversion_rate", "base_metric": "metric_impressions"}
        >>> refs = {"metric_impressions": "$B5", "metric_clicks_cvr": "$D5"}
        >>> _generate_excel_formula("metric_clicks", config, refs)
        "=IF($D5=0,0,$B5*$D5)"

        >>> # Constant
        >>> config = {"formula_type": "constant", "base_metric": "cost_total"}
        >>> refs = {"metric_reach_const": "$D5"}
        >>> _generate_excel_formula("metric_reach", config, refs)
        "=$D5"

        >>> # Power function
        >>> config = {"formula_type": "power_function", "base_metric": "metric_impressions"}
        >>> refs = {"metric_impressions": "$B5", "metric_custom1_coef": "$D5", "metric_custom1_param1": "$E5"}
        >>> _generate_excel_formula("metric_custom1", config, refs)
        "=IF($B5=0,0,$D5*($B5^$E5))"
    """
    # Determine formula type (default to cost_per_unit)
    if formula_config is None:
        formula_type = "cost_per_unit"
        base_metric = "cost_total"
    else:
        formula_type = formula_config.get("formula_type", "cost_per_unit")
        base_metric = formula_config.get("base_metric", "cost_total")

    # Get base metric column reference
    base_ref = column_refs.get(base_metric)
    if not base_ref:
        # Base metric column doesn't exist - return empty string
        logger.warning(f"Base metric '{base_metric}' not found in column references for {metric_name}")
        return ""

    # Generate formula based on type
    if formula_type == "cost_per_unit":
        # Formula: metric = base / coefficient
        # Excel: =IF(coefficient=0, 0, base/coefficient)
        coef_field = f"{metric_name}_cpu"
        coef_ref = column_refs.get(coef_field, "")

        if not coef_ref:
            return ""

        # Special case for impressions (multiply by 1000 for CPM)
        if metric_name == "metric_impressions":
            formula = f"=IF({coef_ref}=0,0,{base_ref}/{coef_ref}*1000)"
        else:
            formula = f"=IF({coef_ref}=0,0,{base_ref}/{coef_ref})"

    elif formula_type == "conversion_rate":
        # Formula: metric = base * coefficient
        # Excel: =IF(coefficient=0, 0, base*coefficient)
        coef_field = f"{metric_name}_cvr"
        coef_ref = column_refs.get(coef_field, "")

        if not coef_ref:
            return ""

        formula = f"=IF({coef_ref}=0,0,{base_ref}*{coef_ref})"

    elif formula_type == "constant":
        # Formula: metric = coefficient (constant value)
        # Excel: ={coefficient_value} directly (no column reference)
        coefficient = _populate_coefficient_column(metric_name, line_item, formula_config)

        if coefficient is None or coefficient == 0:
            formula = "=0"
        else:
            # Format the coefficient value in the formula
            formula = f"={float(coefficient)}"

    elif formula_type == "power_function":
        # Formula: metric = coefficient * (base ^ parameter1)
        # Excel: =IF(base=0, 0, coefficient*(base^parameter1))
        coef_field = f"{metric_name}_coef"
        param1_field = f"{metric_name}_param1"

        coef_ref = column_refs.get(coef_field, "")
        param1_ref = column_refs.get(param1_field, "")

        if not coef_ref or not param1_ref:
            return ""

        formula = f"=IF({base_ref}=0,0,{coef_ref}*({base_ref}^{param1_ref}))"

    else:
        # Unknown formula type
        logger.warning(f"Unknown formula type '{formula_type}' for {metric_name}")
        return ""

    return formula


def _create_v3_styles(workbook: Workbook) -> None:
    """
    Create styles for v3.0 Excel export including formula column styling.

    Args:
        workbook: The workbook to add styles to
    """
    # Header style
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, size=12, color="FFFFFF")
    header_style.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    workbook.add_named_style(header_style)

    # Data style
    data_style = NamedStyle(name="data_style")
    data_style.font = Font(size=11)
    data_style.alignment = Alignment(vertical="center")
    workbook.add_named_style(data_style)

    # Currency style
    currency_style = NamedStyle(name="currency_style")
    currency_style.font = Font(size=11)
    currency_style.number_format = "$#,##0.00"
    currency_style.alignment = Alignment(horizontal="right", vertical="center")
    workbook.add_named_style(currency_style)

    # Date style
    date_style = NamedStyle(name="date_style")
    date_style.font = Font(size=11)
    date_style.number_format = "YYYY-MM-DD"
    date_style.alignment = Alignment(horizontal="center", vertical="center")
    workbook.add_named_style(date_style)

    # Dictionary header style (for custom fields configuration)
    dict_header_style = NamedStyle(name="dict_header_style")
    dict_header_style.font = Font(bold=True, size=11, color="FFFFFF")
    dict_header_style.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    dict_header_style.alignment = Alignment(horizontal="center", vertical="center")
    workbook.add_named_style(dict_header_style)

    # Grey font style for formula columns
    grey_font_style = NamedStyle(name="grey_font_style")
    grey_font_style.font = Font(size=11, color="808080")  # Grey color
    grey_font_style.alignment = Alignment(vertical="center")
    workbook.add_named_style(grey_font_style)


def _populate_v3_workbook(workbook: Workbook, media_plan: "MediaPlan", include_documentation: bool) -> None:
    """
    Populate a workbook with v3.0 schema data.

    Args:
        workbook: The workbook to populate
        media_plan: MediaPlan object (v3.0 schema)
        include_documentation: Whether to include a documentation sheet
    """
    meta = media_plan.meta
    campaign = media_plan.campaign
    line_items = media_plan.lineitems  # List[LineItem] objects
    dictionary = media_plan.dictionary  # Dictionary object

    # Populate all sheets
    _populate_v3_metadata_sheet(workbook["Metadata"], media_plan)
    _populate_v3_campaign_sheet(workbook["Campaign"], campaign)
    _populate_v3_target_audiences_sheet(workbook["Target Audiences"], campaign)  # NEW for v3.0
    _populate_v3_target_locations_sheet(workbook["Target Locations"], campaign)  # NEW for v3.0
    _populate_v3_lineitems_sheet(workbook["Line Items"], line_items, dictionary)  # Pass dictionary for formula-aware export
    _populate_v3_dictionary_sheet(workbook["Dictionary"], dictionary)

    # Populate documentation sheet if needed
    if include_documentation:
        _populate_v3_documentation_sheet(workbook["Documentation"])
    elif "Documentation" in workbook.sheetnames:
        workbook.remove(workbook["Documentation"])


def _populate_v3_metadata_sheet(sheet, media_plan: "MediaPlan") -> None:
    """
    Populate the metadata sheet with v3.0 schema information.

    Args:
        sheet: The worksheet to populate
        media_plan: MediaPlan object
    """
    meta = media_plan.meta

    # Set column widths
    sheet.column_dimensions["A"].width = 22  # Field
    sheet.column_dimensions["B"].width = 40  # Value
    sheet.column_dimensions["C"].width = 10  # Required
    sheet.column_dimensions["D"].width = 60  # Comment

    # Add title
    sheet['A1'] = "Media Plan Metadata (v3.0)"
    sheet['A1'].style = "header_style"
    sheet.merge_cells('A1:D1')

    # Freeze panes after row 1
    sheet.freeze_panes = 'A2'

    # Define metadata fields with required status and descriptions from schema
    # Format: (label, field_key, is_required, description)
    metadata_fields = [
        ("Schema Version:", "schema_version", True, "Version of the media plan schema used (e.g., '3.0')"),
        ("Media Plan ID:", "id", True, "Unique identifier for this media plan document"),
        ("Media Plan Name:", "name", False, "Human-readable name for the media plan"),
        ("Created By Name:", "created_by_name", True, "Full name of the user who created this media plan"),
        ("Created By ID:", "created_by_id", False, "Unique identifier of the user who created this media plan"),
        ("Created At:", "created_at", True, "Timestamp when this media plan was created in ISO 8601 format"),
        ("Is Current:", "is_current", False, "Whether this is the current/active version of the media plan"),
        ("Is Archived:", "is_archived", False, "Whether this media plan has been archived"),
        ("Parent ID:", "parent_id", False, "Identifier of the parent media plan if this is a revision or copy"),
        ("Comments:", "comments", False, "General comments or notes about this media plan"),
    ]

    # Add metadata fields
    row = 2
    for label, field_key, is_required, description in metadata_fields:
        sheet[f'A{row}'] = label
        # Special handling for schema_version
        if field_key == "schema_version":
            sheet[f'B{row}'] = "3.0"
        else:
            sheet[f'B{row}'] = getattr(meta, field_key, "")
        sheet[f'C{row}'] = "TRUE" if is_required else ""
        sheet[f'D{row}'] = description
        row += 1

    # Add custom dimension fields (v3.0)
    for i in range(1, 6):
        sheet[f'A{row}'] = f"Dim Custom {i}:"
        sheet[f'B{row}'] = getattr(meta, f"dim_custom{i}", "")
        sheet[f'C{row}'] = ""
        sheet[f'D{row}'] = f"Custom dimension field {i} - configuration defined in dictionary schema"
        row += 1

    # Add custom properties field (v3.0)
    sheet[f'A{row}'] = "Custom Properties:"
    custom_props = getattr(meta, "custom_properties", None)
    if custom_props:
        import json
        sheet[f'B{row}'] = json.dumps(custom_props)
    else:
        sheet[f'B{row}'] = ""
    sheet[f'C{row}'] = ""
    sheet[f'D{row}'] = "Extensible JSON dictionary for storing custom metadata, settings, or metrics that don't fit elsewhere in the schema"
    row += 1

    # Add export date (not a schema field)
    sheet[f'A{row}'] = "Export Date:"
    sheet[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet[f'C{row}'] = ""
    sheet[f'D{row}'] = "Timestamp when this Excel file was exported"


def _populate_v3_campaign_sheet(sheet, campaign) -> None:
    """
    Populate the campaign sheet with v3.0 schema information.

    Args:
        sheet: The worksheet to populate
        campaign: Campaign object
    """
    import json

    # Set column widths
    sheet.column_dimensions["A"].width = 25  # Field
    sheet.column_dimensions["B"].width = 40  # Value
    sheet.column_dimensions["C"].width = 10  # Required
    sheet.column_dimensions["D"].width = 60  # Comment

    # Add title
    sheet['A1'] = "Campaign Information (v3.0)"
    sheet['A1'].style = "header_style"
    sheet.merge_cells('A1:D1')

    # Freeze panes after row 1
    sheet.freeze_panes = 'A2'

    # Define campaign fields with required status and descriptions from schema
    # Format: (label, field_key, is_required, description, apply_style)
    campaign_fields = [
        ("Campaign ID:", "id", True, "Unique identifier for the campaign", None),
        ("Campaign Name:", "name", True, "Human-readable name for the campaign", None),
        ("Objective:", "objective", False, "Primary marketing objective for the campaign (e.g., awareness, conversion, engagement)", None),
        ("Start Date:", "start_date", True, "Campaign start date in YYYY-MM-DD format", "date_style"),
        ("End Date:", "end_date", True, "Campaign end date in YYYY-MM-DD format", "date_style"),
        ("Budget Total:", "budget_total", True, "Total budget allocated for the campaign in the base currency", "currency_style"),
        ("Budget Currency:", "budget_currency", False, "Currency in which the budget of this campaign is expressed", None),
        ("Agency ID:", "agency_id", False, "Unique identifier for the agency managing the campaign", None),
        ("Agency Name:", "agency_name", False, "Name of the agency managing the campaign", None),
        ("Advertiser ID:", "advertiser_id", False, "Unique identifier for the advertiser/client", None),
        ("Advertiser Name:", "advertiser_name", False, "Name of the advertiser/client organization", None),
        ("Product ID:", "product_id", False, "Unique identifier for the product being advertised", None),
        ("Product Name:", "product_name", False, "Name of the product being advertised", None),
        ("Product Description:", "product_description", False, "Detailed description of the product being advertised", None),
        ("Campaign Type ID:", "campaign_type_id", False, "Unique identifier for the campaign type classification", None),
        ("Campaign Type Name:", "campaign_type_name", False, "Name of the campaign type (e.g., Brand Awareness, Performance, Retargeting)", None),
        ("Workflow Status ID:", "workflow_status_id", False, "Unique identifier for the workflow status", None),
        ("Workflow Status Name:", "workflow_status_name", False, "Name of the current workflow status", None),
    ]

    # Add campaign fields
    row = 2
    for label, field_key, is_required, description, style in campaign_fields:
        sheet[f'A{row}'] = label
        sheet[f'B{row}'] = getattr(campaign,field_key, "")
        if style:
            sheet[f'B{row}'].style = style
        sheet[f'C{row}'] = "TRUE" if is_required else ""
        sheet[f'D{row}'] = description
        row += 1

    # Add KPI fields (v3.0) - alternating name/value pairs
    for i in range(1, 6):
        sheet[f'A{row}'] = f"KPI Name {i}:"
        sheet[f'B{row}'] = getattr(campaign,f"kpi_name{i}", "")
        sheet[f'C{row}'] = ""
        sheet[f'D{row}'] = f"Name of key performance indicator {i}"
        row += 1

        sheet[f'A{row}'] = f"KPI Value {i}:"
        kpi_value = getattr(campaign,f"kpi_value{i}")
        sheet[f'B{row}'] = kpi_value if kpi_value is not None else ""
        sheet[f'C{row}'] = ""
        sheet[f'D{row}'] = f"Target value or goal for key performance indicator {i}"
        row += 1

    # Add custom dimension fields (v3.0)
    for i in range(1, 6):
        sheet[f'A{row}'] = f"Dim Custom {i}:"
        sheet[f'B{row}'] = getattr(campaign,f"dim_custom{i}", "")
        sheet[f'C{row}'] = ""
        sheet[f'D{row}'] = f"Custom dimension field {i} - configuration defined in dictionary schema"
        row += 1

    # Add custom properties field (v3.0)
    sheet[f'A{row}'] = "Custom Properties:"
    custom_props = getattr(campaign,"custom_properties")
    if custom_props:
        sheet[f'B{row}'] = json.dumps(custom_props)
    else:
        sheet[f'B{row}'] = ""
    sheet[f'C{row}'] = ""
    sheet[f'D{row}'] = "Extensible JSON dictionary for storing custom metadata, settings, or metrics that don't fit elsewhere in the schema"


def _populate_v3_target_audiences_sheet(sheet, campaign) -> None:
    """
    Populate the Target Audiences sheet with v3.0 schema data (NEW for v3.0).

    Args:
        sheet: The worksheet to populate
        campaign: The campaign data containing target_audiences array
    """
    # Get target audiences from campaign
    target_audiences = getattr(campaign, "target_audiences", [])
    # Handle None value
    if target_audiences is None:
        target_audiences = []

    # Set column widths
    sheet.column_dimensions["A"].width = 20  # Name
    sheet.column_dimensions["B"].width = 30  # Description
    sheet.column_dimensions["C"].width = 15  # demo_age_start
    sheet.column_dimensions["D"].width = 15  # demo_age_end
    sheet.column_dimensions["E"].width = 15  # demo_gender
    sheet.column_dimensions["F"].width = 25  # demo_attributes
    sheet.column_dimensions["G"].width = 25  # interest_attributes
    sheet.column_dimensions["H"].width = 25  # intent_attributes
    sheet.column_dimensions["I"].width = 25  # purchase_attributes
    sheet.column_dimensions["J"].width = 25  # content_attributes
    sheet.column_dimensions["K"].width = 25  # exclusion_list
    sheet.column_dimensions["L"].width = 20  # extension_approach
    sheet.column_dimensions["M"].width = 15  # population_size

    # Freeze panes at C2 (freeze first two columns)
    sheet.freeze_panes = 'C2'

    # Add headers
    headers = [
        "Name", "Description", "Demo Age Start", "Demo Age End", "Demo Gender",
        "Demo Attributes", "Interest Attributes", "Intent Attributes",
        "Purchase Attributes", "Content Attributes", "Exclusion List",
        "Extension Approach", "Population Size"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.style = "header_style"

    # Add audience data (one row per audience)
    for row_idx, audience in enumerate(target_audiences, 2):
        sheet.cell(row=row_idx, column=1, value=getattr(audience, "name", ""))
        sheet.cell(row=row_idx, column=2, value=getattr(audience, "description", ""))
        sheet.cell(row=row_idx, column=3, value=getattr(audience, "demo_age_start", None))
        sheet.cell(row=row_idx, column=4, value=getattr(audience, "demo_age_end", None))
        sheet.cell(row=row_idx, column=5, value=getattr(audience, "demo_gender", ""))
        sheet.cell(row=row_idx, column=6, value=getattr(audience, "demo_attributes", ""))
        sheet.cell(row=row_idx, column=7, value=getattr(audience, "interest_attributes", ""))
        sheet.cell(row=row_idx, column=8, value=getattr(audience, "intent_attributes", ""))
        sheet.cell(row=row_idx, column=9, value=getattr(audience, "purchase_attributes", ""))
        sheet.cell(row=row_idx, column=10, value=getattr(audience, "content_attributes", ""))
        sheet.cell(row=row_idx, column=11, value=getattr(audience, "exclusion_list", ""))
        sheet.cell(row=row_idx, column=12, value=getattr(audience, "extension_approach", ""))
        sheet.cell(row=row_idx, column=13, value=getattr(audience, "population_size", None))

    logger.info(f"Exported {len(target_audiences)} target audience(s) to Target Audiences sheet")


def _populate_v3_target_locations_sheet(sheet, campaign) -> None:
    """
    Populate the Target Locations sheet with v3.0 schema data (NEW for v3.0).

    Args:
        sheet: The worksheet to populate
        campaign: The campaign data containing target_locations array
    """
    import json

    # Get target locations from campaign
    target_locations = getattr(campaign, "target_locations", [])
    # Handle None value
    if target_locations is None:
        target_locations = []

    # Set column widths
    sheet.column_dimensions["A"].width = 25  # Name
    sheet.column_dimensions["B"].width = 30  # Description
    sheet.column_dimensions["C"].width = 15  # location_type
    sheet.column_dimensions["D"].width = 40  # location_list
    sheet.column_dimensions["E"].width = 15  # exclusion_type
    sheet.column_dimensions["F"].width = 40  # exclusion_list
    sheet.column_dimensions["G"].width = 20  # population_percent

    # Freeze panes at C2 (freeze first two columns)
    sheet.freeze_panes = 'C2'

    # Add headers
    headers = [
        "Name", "Description", "Location Type", "Location List",
        "Exclusion Type", "Exclusion List", "Population Percent"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.style = "header_style"

    # Add location data (one row per location)
    for row_idx, location in enumerate(target_locations, 2):
        sheet.cell(row=row_idx, column=1, value=getattr(location, "name", ""))
        sheet.cell(row=row_idx, column=2, value=getattr(location, "description", ""))
        sheet.cell(row=row_idx, column=3, value=getattr(location, "location_type", ""))

        # location_list is an array - convert to comma-separated string or JSON
        location_list = getattr(location, "location_list", [])
        if location_list:
            if isinstance(location_list, list):
                sheet.cell(row=row_idx, column=4, value=", ".join(location_list))
            else:
                sheet.cell(row=row_idx, column=4, value=location_list)

        sheet.cell(row=row_idx, column=5, value=getattr(location, "exclusion_type", ""))

        # exclusion_list is an array - convert to comma-separated string or JSON
        exclusion_list = getattr(location, "exclusion_list", [])
        if exclusion_list:
            if isinstance(exclusion_list, list):
                sheet.cell(row=row_idx, column=6, value=", ".join(exclusion_list))
            else:
                sheet.cell(row=row_idx, column=6, value=exclusion_list)

        sheet.cell(row=row_idx, column=7, value=getattr(location, "population_percent", None))

    logger.info(f"Exported {len(target_locations)} target location(s) to Target Locations sheet")


def _populate_v3_lineitems_sheet(sheet, line_items: List["LineItem"], dictionary: "Dictionary") -> None:
    """
    Populate the line items sheet with v3.0 schema data including formula-aware calculated columns.

    This function creates Excel formulas that match the formula_type and base_metric
    configurations in the Dictionary. For metrics with no Dictionary configuration,
    defaults to cost_per_unit formulas.

    Args:
        sheet: The worksheet to populate
        line_items: List of LineItem objects
        dictionary: Dictionary object with formula definitions
    """
    # Determine which fields are actually present in line items
    fields_present = set()
    for line_item in line_items:
        # For Pydantic models, use model_dump(exclude_unset=True) to only include fields that were explicitly set
        # This matches the old dict behavior where only present fields were included
        fields_present.update(line_item.model_dump(exclude_unset=True).keys())

    # Define base field order for v3.0 schema
    base_field_order = [
        # Required fields
        ("id", "ID"),
        ("name", "Name"),
        ("start_date", "Start Date"),
        ("end_date", "End Date"),
        ("cost_total", "Cost Total"),

        # Channel-related fields
        ("channel", "Channel"),
        ("channel_custom", "Channel Custom"),
        ("vehicle", "Vehicle"),
        ("vehicle_custom", "Vehicle Custom"),
        ("partner", "Partner"),
        ("partner_custom", "Partner Custom"),
        ("media_product", "Media Product"),
        ("media_product_custom", "Media Product Custom"),

        # Location fields
        ("location_type", "Location Type"),
        ("location_name", "Location Name"),

        # Target/format fields
        ("target_audience", "Target Audience"),
        ("adformat", "Ad Format"),
        ("adformat_custom", "Ad Format Custom"),

        # KPI fields
        ("kpi", "KPI"),
        ("kpi_custom", "KPI Custom"),
        ("kpi_value", "KPI Value"),

        # Dayparts and inventory fields
        ("dayparts", "Dayparts"),
        ("dayparts_custom", "Dayparts Custom"),
        ("inventory", "Inventory"),
        ("inventory_custom", "Inventory Custom"),

        # NEW v3.0: Buy fields
        ("buy_type", "Buy Type"),
        ("buy_commitment", "Buy Commitment"),

        # NEW v3.0: Aggregation fields
        ("is_aggregate", "Is Aggregate"),
        ("aggregation_level", "Aggregation Level"),

        # NEW v3.0: Cost metadata and constraint fields (no % columns needed)
        ("cost_currency", "Cost Currency"),
        ("cost_currency_exchange_rate", "Cost Currency Exchange Rate"),
        ("cost_minimum", "Cost Minimum"),
        ("cost_maximum", "Cost Maximum"),
    ]

    # Add custom dimension fields
    for i in range(1, 11):
        base_field_order.append((f"dim_custom{i}", f"Dim Custom {i}"))

    # Cost fields - we'll insert calculated columns here (% columns for breakdown components)
    cost_fields = [
        ("cost_media", "Cost Media"),
        ("cost_buying", "Cost Buying"),
        ("cost_platform", "Cost Platform"),
        ("cost_data", "Cost Data"),
        ("cost_creative", "Cost Creative"),
    ]

    # Add custom cost fields
    for i in range(1, 11):
        cost_fields.append((f"cost_custom{i}", f"Cost Custom {i}"))

    # Performance metric fields - we'll insert calculated columns here
    performance_fields = [
        ("metric_impressions", "Impressions"),
        ("metric_clicks", "Clicks"),
        ("metric_views", "Views"),
        ("metric_engagements", "Engagements"),
        ("metric_followers", "Followers"),
        ("metric_visits", "Visits"),
        ("metric_leads", "Leads"),
        ("metric_sales", "Sales"),
        ("metric_add_to_cart", "Add to Cart"),
        ("metric_app_install", "App Install"),
        ("metric_application_start", "Application Start"),
        ("metric_application_complete", "Application Complete"),
        ("metric_contact_us", "Contact Us"),
        ("metric_download", "Download"),
        ("metric_signup", "Signup"),
        # NEW v3.0: Additional metrics
        ("metric_view_starts", "View Starts"),
        ("metric_view_completions", "View Completions"),
        ("metric_reach", "Reach"),
        ("metric_units", "Units"),
        ("metric_impression_share", "Impression Share"),
        ("metric_page_views", "Page Views"),
        ("metric_likes", "Likes"),
        ("metric_shares", "Shares"),
        ("metric_comments", "Comments"),
        ("metric_conversions", "Conversions"),
    ]

    # Add custom metric fields
    for i in range(1, 11):
        performance_fields.append((f"metric_custom{i}", f"Metric Custom {i}"))

    # Determine which cost and performance metrics are present
    present_cost_fields = []
    present_performance_fields = []

    # Check cost fields (excluding cost_total and cost_currency which don't need calculated columns)
    for field_name, header_name in cost_fields:
        if field_name != "cost_currency" and (field_name in fields_present):
            present_cost_fields.append((field_name, header_name))

    # Check performance fields (excluding max daily fields and audience size which are limits, not performance)
    for field_name, header_name in performance_fields:
        if field_name in fields_present:
            present_performance_fields.append((field_name, header_name))

    # Build dynamic field order with calculated columns
    dynamic_field_order = []

    # Add base fields (up to custom dimensions)
    for field_name, header_name in base_field_order:
        if field_name in fields_present or field_name in {"id", "name", "start_date", "end_date", "cost_total"}:
            dynamic_field_order.append((field_name, header_name, "base"))

    # Add cost fields with calculated columns
    for field_name, header_name in present_cost_fields:
        # Add percentage column before actual cost column
        calc_field_name = f"{field_name}_pct"
        calc_header_name = f"{header_name} %"
        dynamic_field_order.append((calc_field_name, calc_header_name, "calculated"))

        # Add actual cost column
        dynamic_field_order.append((field_name, header_name, "formula"))

    # FORMULA-AWARE: Add missing base metrics if needed (Design Decision 1)
    missing_base_metrics = _get_missing_base_metrics(present_performance_fields, dictionary)
    for missing_field, missing_header in missing_base_metrics:
        # Add as base column (will be populated with 0 values)
        # Note: Do NOT add to present_performance_fields to avoid duplicate processing
        dynamic_field_order.append((missing_field, missing_header, "base"))

    # FORMULA-AWARE: Add performance fields with calculated columns based on Dictionary config
    for field_name, header_name in present_performance_fields:
        # Get formula configuration from Dictionary
        formula_config = _get_formula_config(field_name, dictionary)

        # Determine which calculated columns to create based on formula_type
        columns_to_add = _determine_calculated_columns(field_name, header_name, formula_config)

        # Add all columns (calculated + metric)
        for col_field, col_header, col_type in columns_to_add:
            dynamic_field_order.append((col_field, col_header, col_type))

    # Add remaining metric fields that don't get calculated columns
    remaining_metrics = [
        ("metric_max_daily_spend", "Max Daily Spend"),
        ("metric_max_daily_impressions", "Max Daily Impressions"),
        ("metric_audience_size", "Audience Size"),
    ]

    for field_name, header_name in remaining_metrics:
        if field_name in fields_present:
            dynamic_field_order.append((field_name, header_name, "base"))

    # NEW v3.0: Add metric_formulas and custom_properties at the end
    if "metric_formulas" in fields_present:
        dynamic_field_order.append(("metric_formulas", "Metric Formulas (JSON)", "json"))

    if "custom_properties" in fields_present:
        dynamic_field_order.append(("custom_properties", "Custom Properties (JSON)", "json"))

    # Set column widths based on field type
    for col_idx, (field_name, header_name, field_type) in enumerate(dynamic_field_order, 1):
        width = 15
        if field_name == "cost_total":
            width = 22.5  # 50% wider than default for Cost Total column
        elif field_name in ["name", "media_product", "media_product_custom", "target_audience"]:
            width = 25
        elif field_name in ["partner", "partner_custom", "vehicle", "vehicle_custom"]:
            width = 20
        elif field_type == "calculated":
            width = 18  # Slightly wider for calculated columns
        elif field_type == "json":
            width = 40  # Wider for JSON data
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    # Add headers with appropriate styling
    for col_idx, (field_name, header_name, field_type) in enumerate(dynamic_field_order, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header_name)
        cell.style = "header_style"

    # Freeze panes at C2 (freeze first two columns: ID and Name)
    sheet.freeze_panes = 'C2'

    # Add line item data with calculated values and formulas
    for row_idx, line_item in enumerate(line_items, 2):
        cost_total_col = None

        # First pass: populate base values and find cost_total column
        for col_idx, (field_name, header_name, field_type) in enumerate(dynamic_field_order, 1):
            if field_name == "cost_total":
                cost_total_col = col_idx

            if field_type == "base":
                # Access LineItem object attribute
                value = getattr(line_item, field_name, None)
                if value is None:
                    continue

                # Apply appropriate formatting based on field type
                if field_name in ["start_date", "end_date"]:
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.style = "date_style"
                elif field_name.startswith("cost") or field_name.startswith("metric") or field_name == "cost_total":
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if field_name.startswith("cost"):
                        cell.style = "currency_style"
                    elif field_name.startswith("metric"):
                        # Add thousand comma separator for all metrics
                        cell.number_format = '#,##0'
                else:
                    sheet.cell(row=row_idx, column=col_idx, value=value)

        # Second pass: populate calculated columns and formulas (FORMULA-AWARE)
        cost_total_value = getattr(line_item, "cost_total", 0)
        cost_total_cell_ref = f"{get_column_letter(cost_total_col)}{row_idx}" if cost_total_col else "0"

        # Build column reference map for formula generation
        column_refs = {}
        for idx, (fname, _, _) in enumerate(dynamic_field_order, 1):
            column_refs[fname] = f"${get_column_letter(idx)}{row_idx}"

        for col_idx, (field_name, header_name, field_type) in enumerate(dynamic_field_order, 1):
            if field_type == "calculated":
                if field_name.endswith("_pct"):
                    # Cost percentage calculation (unchanged)
                    base_field = field_name.replace("_pct", "")
                    cost_value = getattr(line_item, base_field, 0)
                    # Handle None values
                    if cost_value is None:
                        cost_value = 0

                    if cost_total_value and cost_total_value != 0:
                        percentage = (cost_value / cost_total_value)
                    else:
                        percentage = 0

                    cell = sheet.cell(row=row_idx, column=col_idx, value=percentage)
                    cell.number_format = '0.0%'  # Percentage format with 1 decimal place

                elif field_name.endswith(("_cpu", "_cvr", "_coef")):
                    # FORMULA-AWARE: Coefficient column (CPU, CVR, or Power Coefficient)
                    # Extract metric name from calculated field name
                    if field_name.endswith("_cpu"):
                        metric_name = field_name.replace("_cpu", "")
                    elif field_name.endswith("_cvr"):
                        metric_name = field_name.replace("_cvr", "")
                    elif field_name.endswith("_coef"):
                        metric_name = field_name.replace("_coef", "")

                    # Get formula config and calculate coefficient
                    formula_config = _get_formula_config(metric_name, dictionary)
                    coefficient = _populate_coefficient_column(metric_name, line_item, formula_config)

                    if coefficient is not None:
                        cell = sheet.cell(row=row_idx, column=col_idx, value=float(coefficient))
                        # Apply appropriate format based on formula type
                        if field_name.endswith("_cvr"):
                            cell.number_format = '0.00%'  # Percentage format for conversion rates
                        elif field_name.endswith("_cpu"):
                            cell.number_format = '$#,##0.00'  # Currency format for cost per unit
                        else:
                            cell.number_format = '0.0000'  # 4 decimal places (Decision 4)
                    else:
                        cell = sheet.cell(row=row_idx, column=col_idx, value=0)
                        # Apply appropriate format based on formula type
                        if field_name.endswith("_cvr"):
                            cell.number_format = '0.00%'
                        elif field_name.endswith("_cpu"):
                            cell.number_format = '$#,##0.00'  # Currency format for cost per unit
                        else:
                            cell.number_format = '0.0000'

                elif field_name.endswith("_param1"):
                    # FORMULA-AWARE: Parameter1 column for power_function
                    metric_name = field_name.replace("_param1", "")
                    parameter1 = _get_parameter1_value(metric_name, line_item)

                    cell = sheet.cell(row=row_idx, column=col_idx, value=float(parameter1))
                    cell.number_format = '0.0000'  # 4 decimal places

            elif field_type == "formula":
                if field_name.startswith("cost_") and field_name != "cost_total" and field_name != "cost_currency":
                    # Cost field formula: cost_total * percentage (unchanged)
                    pct_col_idx = None
                    for idx, (fname, _, ftype) in enumerate(dynamic_field_order, 1):
                        if fname == f"{field_name}_pct":
                            pct_col_idx = idx
                            break

                    if pct_col_idx:
                        pct_cell_ref = f"{get_column_letter(pct_col_idx)}{row_idx}"
                        formula = f"={cost_total_cell_ref}*{pct_cell_ref}"
                        cell = sheet.cell(row=row_idx, column=col_idx, value=formula)
                        cell.style = "currency_style"
                        # Apply grey font formatting for formula columns
                        cell.font = Font(color="808080")  # Grey color

                elif field_name.startswith("metric_"):
                    # FORMULA-AWARE: Performance metric formula
                    formula_config = _get_formula_config(field_name, dictionary)
                    formula = _generate_excel_formula(field_name, formula_config, column_refs, line_item)

                    if formula:
                        cell = sheet.cell(row=row_idx, column=col_idx, value=formula)
                        # Apply grey font formatting and comma separator for formula columns
                        cell.font = Font(color="808080")  # Grey color
                        cell.number_format = "#,##0"  # Thousand comma separator, no decimals
                    else:
                        # Formula generation failed - show metric value directly
                        metric_value = getattr(line_item, field_name, 0)
                        cell = sheet.cell(row=row_idx, column=col_idx, value=metric_value)
                        cell.number_format = "#,##0"

            elif field_type == "json":
                # NEW v3.0: Handle JSON fields (metric_formulas, custom_properties)
                value = getattr(line_item, field_name, None)
                if value is not None:
                    import json
                    # Convert to dict if it's a Pydantic model, then to JSON string
                    if hasattr(value, 'model_dump'):
                        # Single Pydantic model - convert to dict first
                        json_string = json.dumps(value.model_dump(exclude_none=True))
                    elif isinstance(value, dict):
                        # Dict that may contain Pydantic models as values (e.g., Dict[str, MetricFormula])
                        converted_dict = {}
                        for key, val in value.items():
                            if hasattr(val, 'model_dump'):
                                # Convert Pydantic model to dict
                                converted_dict[key] = val.model_dump(exclude_none=True)
                            else:
                                converted_dict[key] = val
                        json_string = json.dumps(converted_dict)
                    else:
                        json_string = str(value)
                    sheet.cell(row=row_idx, column=col_idx, value=json_string)

    logger.info(
        f"Created {len(dynamic_field_order)} columns with {len(present_cost_fields)} cost calculations and {len(present_performance_fields)} performance calculations")


def _get_config_value(config, field_name: str, default=""):
    """
    Helper to get value from config which can be either Pydantic object or dict.

    Args:
        config: Either a Pydantic model (CustomFieldConfig, StandardMetricConfig) or plain dict
        field_name: Name of the field to retrieve
        default: Default value if not found

    Returns:
        The field value or default
    """
    if hasattr(config, field_name):
        # Pydantic object
        return getattr(config, field_name, default)
    else:
        # Plain dict
        return config.get(field_name, default)


def _populate_v3_dictionary_sheet(sheet, dictionary: "Dictionary") -> None:
    """
    Populate the dictionary configuration sheet with v3.0 structure.

    v3.0 Dictionary includes:
    - Meta, Campaign, and LineItem custom dimensions (status, caption)
    - Standard metrics (formula_type, base_metric)
    - Custom metrics (status, caption, formula_type, base_metric)
    - Custom costs (status, caption)

    Args:
        sheet: The worksheet to populate
        dictionary: The dictionary configuration data
    """
    # Set column widths for 7 columns
    sheet.column_dimensions["A"].width = 25  # Field Name
    sheet.column_dimensions["B"].width = 20  # Field Type
    sheet.column_dimensions["C"].width = 25  # Column Name
    sheet.column_dimensions["D"].width = 35  # Caption
    sheet.column_dimensions["E"].width = 12  # Status
    sheet.column_dimensions["F"].width = 20  # Formula Type
    sheet.column_dimensions["G"].width = 20  # Base Metric

    # Add title
    sheet['A1'] = "Dictionary Configuration (v3.0)"
    sheet['A1'].style = "header_style"
    sheet.merge_cells('A1:G1')

    # Add headers
    row = 2
    sheet[f'A{row}'] = "Field Name"
    sheet[f'B{row}'] = "Field Type"
    sheet[f'C{row}'] = "Column Name"
    sheet[f'D{row}'] = "Caption"
    sheet[f'E{row}'] = "Status"
    sheet[f'F{row}'] = "Formula Type"
    sheet[f'G{row}'] = "Base Metric"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        sheet[f'{col}{row}'].style = "dict_header_style"

    # Freeze panes at row 3 (after header rows)
    sheet.freeze_panes = 'A3'

    # Add all fields with their current configuration
    row = 3

    # Section 1: Meta custom dimensions (5 fields)
    meta_custom_dimensions = getattr(dictionary,"meta_custom_dimensions", {})
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for i in range(1, 6):
        field_name = f"dim_custom{i}"
        config = meta_custom_dimensions.get(field_name, {"status": "disabled", "caption": ""})
        column_name = f"Meta Dim Custom {i}"

        sheet[f'A{row}'] = field_name
        sheet[f'B{row}'] = "Meta Dimension"
        sheet[f'C{row}'] = column_name
        sheet[f'D{row}'] = _get_config_value(config, "caption", "")
        sheet[f'E{row}'] = _get_config_value(config, "status", "disabled")
        sheet[f'F{row}'] = ""  # No formula for dimensions
        sheet[f'G{row}'] = ""  # No base metric for dimensions
        # Grey out formula columns for dimensions
        sheet[f'F{row}'].fill = grey_fill
        sheet[f'G{row}'].fill = grey_fill
        row += 1

    # Section 2: Campaign custom dimensions (5 fields)
    campaign_custom_dimensions = getattr(dictionary,"campaign_custom_dimensions", {})
    for i in range(1, 6):
        field_name = f"dim_custom{i}"
        config = campaign_custom_dimensions.get(field_name, {"status": "disabled", "caption": ""})
        column_name = f"Campaign Dim Custom {i}"

        sheet[f'A{row}'] = field_name
        sheet[f'B{row}'] = "Campaign Dimension"
        sheet[f'C{row}'] = column_name
        sheet[f'D{row}'] = _get_config_value(config, "caption", "")
        sheet[f'E{row}'] = _get_config_value(config, "status", "disabled")
        sheet[f'F{row}'] = ""  # No formula for dimensions
        sheet[f'G{row}'] = ""  # No base metric for dimensions
        # Grey out formula columns for dimensions
        sheet[f'F{row}'].fill = grey_fill
        sheet[f'G{row}'].fill = grey_fill
        row += 1

    # Section 3: LineItem custom dimensions (10 fields)
    lineitem_custom_dimensions = getattr(dictionary,"lineitem_custom_dimensions", {})
    for i in range(1, 11):
        field_name = f"dim_custom{i}"
        config = lineitem_custom_dimensions.get(field_name, {"status": "disabled", "caption": ""})
        column_name = f"LineItem Dim Custom {i}"

        sheet[f'A{row}'] = field_name
        sheet[f'B{row}'] = "LineItem Dimension"
        sheet[f'C{row}'] = column_name
        sheet[f'D{row}'] = _get_config_value(config, "caption", "")
        sheet[f'E{row}'] = _get_config_value(config, "status", "disabled")
        sheet[f'F{row}'] = ""  # No formula for dimensions
        sheet[f'G{row}'] = ""  # No base metric for dimensions
        # Grey out formula columns for dimensions
        sheet[f'F{row}'].fill = grey_fill
        sheet[f'G{row}'].fill = grey_fill
        row += 1

    # Section 4: Standard metrics (25 fields) - NEW in v3.0
    standard_metrics = getattr(dictionary,"standard_metrics", {})
    standard_metric_list = [
        ("metric_impressions", "Impressions"),
        ("metric_clicks", "Clicks"),
        ("metric_views", "Views"),
        ("metric_view_starts", "View Starts"),
        ("metric_view_completions", "View Completions"),
        ("metric_reach", "Reach"),
        ("metric_units", "Units"),
        ("metric_impression_share", "Impression Share"),
        ("metric_engagements", "Engagements"),
        ("metric_followers", "Followers"),
        ("metric_visits", "Visits"),
        ("metric_leads", "Leads"),
        ("metric_sales", "Sales"),
        ("metric_add_to_cart", "Add to Cart"),
        ("metric_app_install", "App Install"),
        ("metric_application_start", "Application Start"),
        ("metric_application_complete", "Application Complete"),
        ("metric_contact_us", "Contact Us"),
        ("metric_download", "Download"),
        ("metric_signup", "Signup"),
        ("metric_page_views", "Page Views"),
        ("metric_likes", "Likes"),
        ("metric_shares", "Shares"),
        ("metric_comments", "Comments"),
        ("metric_conversions", "Conversions"),
    ]

    for field_name, column_name in standard_metric_list:
        config = standard_metrics.get(field_name, {})

        sheet[f'A{row}'] = field_name
        sheet[f'B{row}'] = "Standard Metric"
        sheet[f'C{row}'] = column_name
        sheet[f'D{row}'] = ""  # No caption for standard metrics
        sheet[f'E{row}'] = ""  # No status for standard metrics
        sheet[f'F{row}'] = _get_config_value(config, "formula_type", "")
        sheet[f'G{row}'] = _get_config_value(config, "base_metric", "")
        # Grey out Caption and Status columns for standard metrics
        sheet[f'D{row}'].fill = grey_fill
        sheet[f'E{row}'].fill = grey_fill
        row += 1

    # Section 5: Custom metrics (10 fields) - Updated in v3.0 to include formula support
    custom_metrics = getattr(dictionary,"custom_metrics", {})
    for i in range(1, 11):
        field_name = f"metric_custom{i}"
        config = custom_metrics.get(field_name, {"status": "disabled", "caption": ""})
        column_name = f"Metric Custom {i}"

        sheet[f'A{row}'] = field_name
        sheet[f'B{row}'] = "Custom Metric"
        sheet[f'C{row}'] = column_name
        sheet[f'D{row}'] = _get_config_value(config, "caption", "")
        sheet[f'E{row}'] = _get_config_value(config, "status", "disabled")
        sheet[f'F{row}'] = _get_config_value(config, "formula_type", "")
        sheet[f'G{row}'] = _get_config_value(config, "base_metric", "")
        row += 1

    # Section 6: Custom costs (10 fields)
    custom_costs = getattr(dictionary,"custom_costs", {})
    for i in range(1, 11):
        field_name = f"cost_custom{i}"
        config = custom_costs.get(field_name, {"status": "disabled", "caption": ""})
        column_name = f"Cost Custom {i}"

        sheet[f'A{row}'] = field_name
        sheet[f'B{row}'] = "Custom Cost"
        sheet[f'C{row}'] = column_name
        sheet[f'D{row}'] = _get_config_value(config, "caption", "")
        sheet[f'E{row}'] = _get_config_value(config, "status", "disabled")
        sheet[f'F{row}'] = ""  # No formula for costs
        sheet[f'G{row}'] = ""  # No base metric for costs
        # Grey out formula columns for costs
        sheet[f'F{row}'].fill = grey_fill
        sheet[f'G{row}'].fill = grey_fill
        row += 1

    # Add instructions
    row += 2
    sheet[f'A{row}'] = "Instructions:"
    sheet[f'A{row}'].font = Font(bold=True)
    sheet.merge_cells(f'A{row}:G{row}')

    row += 1
    sheet[f'A{row}'] = "- Custom Fields: Set Status to 'enabled' or 'disabled', Caption is required when enabled"
    sheet.merge_cells(f'A{row}:G{row}')

    row += 1
    sheet[f'A{row}'] = "- Standard/Custom Metrics: Optionally specify Formula Type and Base Metric for formula-based calculations"
    sheet.merge_cells(f'A{row}:G{row}')

    row += 1
    sheet[f'A{row}'] = "- Common formula types: 'cost_per_unit', 'conversion_rate', 'constant', 'power_function'"
    sheet.merge_cells(f'A{row}:G{row}')

    row += 1
    sheet[f'A{row}'] = "- Use the 'Column Name' exactly as shown when importing from Excel"
    sheet.merge_cells(f'A{row}:G{row}')


def _populate_v3_documentation_sheet(sheet) -> None:
    """
    Populate the documentation sheet with comprehensive v3.0 schema information.

    Args:
        sheet: The worksheet to populate
    """
    # Set column widths - updated for new Field Name column
    sheet.column_dimensions["A"].width = 25  # Column Name
    sheet.column_dimensions["B"].width = 20  # Field Name (NEW)
    sheet.column_dimensions["C"].width = 15  # Data Type
    sheet.column_dimensions["D"].width = 67.5  # Description (increased by 50% from 45)
    sheet.column_dimensions["E"].width = 12  # Required

    # Add title
    sheet['A1'] = "Media Plan Excel Documentation (v3.0)"
    sheet['A1'].style = "header_style"
    sheet.merge_cells('A1:E1')  # Updated merge range

    # Add documentation content
    row = 2
    sheet[f'A{row}'] = "Schema Version:"
    sheet[f'B{row}'] = "3.0"

    row += 1
    sheet[f'A{row}'] = "Export Date:"
    sheet[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    row += 2
    sheet[f'A{row}'] = "Instructions:"
    sheet[f'B{row}'] = "This Excel file contains a media plan following the Media Plan Open Data Standard v3.0."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "Use the tabs to navigate through different sections of the media plan."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 2
    sheet[f'A{row}'] = "Sheets:"
    sheet[f'B{row}'] = "Metadata: Contains information about the media plan itself."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "Campaign: Contains campaign details and settings."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "Target Audiences: Contains target audience definitions (NEW in v3.0)."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "Target Locations: Contains target location definitions (NEW in v3.0)."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "Line Items: Contains all line items in the campaign."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "Dictionary: Configuration for custom fields."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    # Comprehensive field documentation header
    row += 2
    sheet[f'A{row}'] = "Supported Line Item Columns"
    sheet[f'A{row}'].font = Font(bold=True, size=12)
    sheet.merge_cells(f'A{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'A{row}'] = "Column Name"
    sheet[f'B{row}'] = "Field Name"  # NEW COLUMN
    sheet[f'C{row}'] = "Data Type"
    sheet[f'D{row}'] = "Description"
    sheet[f'E{row}'] = "Required"
    for col in ['A', 'B', 'C', 'D', 'E']:
        sheet[f'{col}{row}'].style = "header_style"

    # Comprehensive field documentation in schema order with field names
    all_fields_documentation = [
        # Required fields (in schema order)
        ("ID", "id", "Text", "Unique identifier for the line item", "TRUE"),
        ("Name", "name", "Text", "Human-readable name for the line item", "TRUE"),
        ("Start Date", "start_date", "Date", "Line item start date in YYYY-MM-DD format", "TRUE"),
        ("End Date", "end_date", "Date", "Line item end date in YYYY-MM-DD format", "TRUE"),
        ("Cost Total", "cost_total", "Number", "Total cost for the line item including all cost components", "TRUE"),

        # Channel-related fields (in schema order)
        ("Channel", "channel", "Text", "Media channel for the line item (e.g., Digital, TV, Radio, Print)", ""),
        ("Channel Custom", "channel_custom", "Text",
         "Custom channel specification when standard channel options don't apply", ""),
        ("Vehicle", "vehicle", "Text", "Media vehicle or platform (e.g., Facebook, Google, CNN, Spotify)", ""),
        ("Vehicle Custom", "vehicle_custom", "Text",
         "Custom vehicle specification when standard vehicle options don't apply", ""),
        ("Partner", "partner", "Text", "Media partner or vendor handling the placement", ""),
        ("Partner Custom", "partner_custom", "Text",
         "Custom partner specification when standard partner options don't apply", ""),
        ("Media Product", "media_product", "Text", "Specific media product or ad unit being purchased", ""),
        ("Media Product Custom", "media_product_custom", "Text",
         "Custom media product specification when standard options don't apply", ""),

        # Location fields (line-item-level targeting, in schema order)
        ("Location Type", "location_type", "Text",
         "Geographic scope type for the line item targeting (Country or State)", ""),
        ("Location Name", "location_name", "Text", "Name of the geographic location being targeted", ""),

        # Target/format fields (in schema order)
        ("Target Audience", "target_audience", "Text", "Description of the target audience for this line item", ""),
        ("Ad Format", "adformat", "Text", "Creative format or ad type (e.g., Banner, Video, Native)", ""),
        ("Ad Format Custom", "adformat_custom", "Text",
         "Custom ad format specification when standard formats don't apply", ""),
        ("KPI", "kpi", "Text", "Primary key performance indicator for the line item", ""),
        ("KPI Custom", "kpi_custom", "Text", "Custom KPI specification when standard KPIs don't apply", ""),
        ("KPI Value", "kpi_value", "Number", "Target value or goal for the primary key performance indicator", ""),

        # Dayparts and inventory fields (in schema order)
        ("Dayparts", "dayparts", "Text", "Time periods when the ad should run (e.g., Primetime, Morning, All Day)",
         ""),
        (
        "Dayparts Custom", "dayparts_custom", "Text", "Custom daypart specification when standard dayparts don't apply",
        ""),
        ("Inventory", "inventory", "Text", "Type of inventory or placement being purchased", ""),
        ("Inventory Custom", "inventory_custom", "Text",
         "Custom inventory specification when standard inventory types don't apply", ""),

        # Buy fields (NEW in v3.0, in schema order)
        ("Buy Type", "buy_type", "Text",
         "Type of media buying arrangement (e.g., Auction, Programmatic Guaranteed, Upfront, Scatter)", ""),
        ("Buy Commitment", "buy_commitment", "Text",
         "Commitment level for the media purchase (e.g., Cancellable, Committed, Non-Cancellable)", ""),

        # Custom dimension fields (dim_custom1-10, in schema order)
        ("Dim Custom 1", "dim_custom1", "Text", "Custom dimension field 1 - configuration defined in dictionary schema",
         ""),
        ("Dim Custom 2", "dim_custom2", "Text", "Custom dimension field 2 - configuration defined in dictionary schema",
         ""),
        ("Dim Custom 3", "dim_custom3", "Text", "Custom dimension field 3 - configuration defined in dictionary schema",
         ""),
        ("Dim Custom 4", "dim_custom4", "Text", "Custom dimension field 4 - configuration defined in dictionary schema",
         ""),
        ("Dim Custom 5", "dim_custom5", "Text", "Custom dimension field 5 - configuration defined in dictionary schema",
         ""),
        ("Dim Custom 6", "dim_custom6", "Text", "Custom dimension field 6 - configuration defined in dictionary schema",
         ""),
        ("Dim Custom 7", "dim_custom7", "Text", "Custom dimension field 7 - configuration defined in dictionary schema",
         ""),
        ("Dim Custom 8", "dim_custom8", "Text", "Custom dimension field 8 - configuration defined in dictionary schema",
         ""),
        ("Dim Custom 9", "dim_custom9", "Text", "Custom dimension field 9 - configuration defined in dictionary schema",
         ""),
        ("Dim Custom 10", "dim_custom10", "Text",
         "Custom dimension field 10 - configuration defined in dictionary schema", ""),

        # Aggregation fields (NEW in v3.0, in schema order)
        ("Is Aggregate", "is_aggregate", "Boolean",
         "Whether this line item contains aggregated values (useful for storing channel-level budgets or campaign-level reach estimates)", ""),
        ("Aggregation Level", "aggregation_level", "Text",
         "Level at which the aggregate is stored (e.g., channel, campaign, vehicle) when is_aggregate is true", ""),

        # Cost fields (in schema order)
        ("Cost Currency", "cost_currency", "Text",
         "Currency code for all cost fields in this line item (e.g., USD, EUR, GBP)", ""),
        ("Cost Currency Exchange Rate", "cost_currency_exchange_rate", "Number",
         "Exchange rate to convert from line item currency to campaign-level currency (useful when line items use different currencies)", ""),
        ("Cost Media", "cost_media", "Number", "Media cost component (working media spend)", ""),
        ("Cost Buying", "cost_buying", "Number", "Media buying/trading cost component", ""),
        ("Cost Platform", "cost_platform", "Number", "Platform or technology cost component", ""),
        ("Cost Data", "cost_data", "Number", "Data cost component (audience data, targeting data, etc.)", ""),
        ("Cost Creative", "cost_creative", "Number", "Creative production and development cost component", ""),

        # Custom cost fields (cost_custom1-10, in schema order)
        ("Cost Custom 1", "cost_custom1", "Number", "Custom cost field 1 - configuration defined in dictionary schema",
         ""),
        ("Cost Custom 2", "cost_custom2", "Number", "Custom cost field 2 - configuration defined in dictionary schema",
         ""),
        ("Cost Custom 3", "cost_custom3", "Number", "Custom cost field 3 - configuration defined in dictionary schema",
         ""),
        ("Cost Custom 4", "cost_custom4", "Number", "Custom cost field 4 - configuration defined in dictionary schema",
         ""),
        ("Cost Custom 5", "cost_custom5", "Number", "Custom cost field 5 - configuration defined in dictionary schema",
         ""),
        ("Cost Custom 6", "cost_custom6", "Number", "Custom cost field 6 - configuration defined in dictionary schema",
         ""),
        ("Cost Custom 7", "cost_custom7", "Number", "Custom cost field 7 - configuration defined in dictionary schema",
         ""),
        ("Cost Custom 8", "cost_custom8", "Number", "Custom cost field 8 - configuration defined in dictionary schema",
         ""),
        ("Cost Custom 9", "cost_custom9", "Number", "Custom cost field 9 - configuration defined in dictionary schema",
         ""),
        ("Cost Custom 10", "cost_custom10", "Number",
         "Custom cost field 10 - configuration defined in dictionary schema", ""),

        # Cost constraints (NEW in v3.0, in schema order)
        ("Cost Minimum", "cost_minimum", "Number",
         "Minimum budget constraint for the line item (user-defined lower bound)", ""),
        ("Cost Maximum", "cost_maximum", "Number",
         "Maximum budget constraint for the line item (user-defined upper bound)", ""),

        # Standard metric fields (in v3.0 schema order)
        ("Impressions", "metric_impressions", "Number", "Number of ad impressions delivered or planned", ""),
        ("Clicks", "metric_clicks", "Number", "Number of clicks on the ad", ""),
        ("Views", "metric_views", "Number", "Number of video views or content views", ""),
        ("View Starts", "metric_view_starts", "Number", "Number of video view starts", ""),
        ("View Completions", "metric_view_completions", "Number", "Number of video view completions", ""),
        ("Reach", "metric_reach", "Number", "Number of unique users reached", ""),
        ("Units", "metric_units", "Number", "Number of units delivered (e.g., TV spots, radio ads, print insertions)", ""),
        ("Impression Share", "metric_impression_share", "Number", "Percentage of total available impressions captured", ""),
        ("Engagements", "metric_engagements", "Number", "Number of user engagements (likes, shares, comments, etc.)",
         ""),
        ("Followers", "metric_followers", "Number", "Number of new followers gained", ""),
        ("Visits", "metric_visits", "Number", "Number of website visits or page visits", ""),
        ("Leads", "metric_leads", "Number", "Number of leads generated", ""),
        ("Sales", "metric_sales", "Number", "Number of sales or purchases", ""),
        ("Add to Cart", "metric_add_to_cart", "Number", "Number of add-to-cart actions", ""),
        ("App Install", "metric_app_install", "Number", "Number of app installations", ""),
        ("Application Start", "metric_application_start", "Number", "Number of application forms started", ""),
        (
        "Application Complete", "metric_application_complete", "Number", "Number of application forms completed", ""),
        ("Contact Us", "metric_contact_us", "Number", "Number of contact form submissions or contact actions", ""),
        ("Download", "metric_download", "Number", "Number of downloads (files, apps, content)", ""),
        ("Signup", "metric_signup", "Number", "Number of signups or registrations", ""),
        ("Page Views", "metric_page_views", "Number", "Number of page views", ""),
        ("Likes", "metric_likes", "Number", "Number of likes or reactions", ""),
        ("Shares", "metric_shares", "Number", "Number of shares or reposts", ""),
        ("Comments", "metric_comments", "Number", "Number of comments", ""),
        ("Conversions", "metric_conversions", "Number", "Number of conversions or goal completions", ""),
        ("Max Daily Spend", "metric_max_daily_spend", "Number", "Maximum daily spend limit for the line item", ""),
        ("Max Daily Impressions", "metric_max_daily_impressions", "Number",
         "Maximum daily impressions limit for the line item", ""),
        ("Audience Size", "metric_audience_size", "Number", "Size of the targetable audience for this line item", ""),

        # Custom metric fields (metric_custom1-10, in schema order)
        ("Metric Custom 1", "metric_custom1", "Number",
         "Custom metric field 1 - configuration defined in dictionary schema", ""),
        ("Metric Custom 2", "metric_custom2", "Number",
         "Custom metric field 2 - configuration defined in dictionary schema", ""),
        ("Metric Custom 3", "metric_custom3", "Number",
         "Custom metric field 3 - configuration defined in dictionary schema", ""),
        ("Metric Custom 4", "metric_custom4", "Number",
         "Custom metric field 4 - configuration defined in dictionary schema", ""),
        ("Metric Custom 5", "metric_custom5", "Number",
         "Custom metric field 5 - configuration defined in dictionary schema", ""),
        ("Metric Custom 6", "metric_custom6", "Number",
         "Custom metric field 6 - configuration defined in dictionary schema", ""),
        ("Metric Custom 7", "metric_custom7", "Number",
         "Custom metric field 7 - configuration defined in dictionary schema", ""),
        ("Metric Custom 8", "metric_custom8", "Number",
         "Custom metric field 8 - configuration defined in dictionary schema", ""),
        ("Metric Custom 9", "metric_custom9", "Number",
         "Custom metric field 9 - configuration defined in dictionary schema", ""),
        ("Metric Custom 10", "metric_custom10", "Number",
         "Custom metric field 10 - configuration defined in dictionary schema", ""),

        # Metric formulas (NEW in v3.0, in schema order)
        ("Metric Formulas (JSON)", "metric_formulas", "JSON",
         "Formula configurations for metrics that use custom calculation formulas. Each metric's formula type and base metric are defined in the dictionary schema.", ""),

        # Custom properties (NEW in v3.0, in schema order)
        ("Custom Properties (JSON)", "custom_properties", "JSON",
         "Extensible JSON dictionary for storing custom metadata, settings, or metrics that don't fit elsewhere in the schema", ""),
    ]

    # Populate the comprehensive field documentation
    for column_name, field_name, data_type, description, required in all_fields_documentation:
        row += 1
        sheet[f'A{row}'] = column_name
        sheet[f'B{row}'] = field_name  # NEW: Schema field name
        sheet[f'C{row}'] = data_type
        sheet[f'D{row}'] = description
        sheet[f'E{row}'] = required

        # Apply alternating row colors for readability
        if row % 2 == 0:
            for col in ['A', 'B', 'C', 'D', 'E']:
                sheet[f'{col}{row}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    # Add additional v2.0 information
    row += 2
    sheet[f'A{row}'] = "Custom Fields:"
    sheet[
        f'B{row}'] = "v2.0 introduces 30 custom fields (10 dimensions, 10 metrics, 10 costs) that can be configured via the Dictionary sheet."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "- Use the Dictionary sheet to enable/disable custom fields and set their captions"
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "- When enabled, custom fields appear as additional columns in the Line Items sheet"
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "- Use the exact 'Column Name' from this table when importing data"
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 2
    sheet[f'A{row}'] = "Field Mapping:"
    sheet[
        f'B{row}'] = "The 'Field Name' column shows the underlying schema field name that corresponds to each Excel column."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 1
    sheet[f'B{row}'] = "This mapping is useful for developers working with the MediaPlan schema programmatically."
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range

    row += 2
    sheet[f'A{row}'] = "Support:"
    sheet[f'B{row}'] = "For more information, see: https://github.com/planmatic/mediaplanschema"
    sheet.merge_cells(f'B{row}:E{row}')  # Updated merge range


def _add_v3_validation_and_formatting(workbook: Workbook) -> None:
    """
    Add data validation and formatting for v3.0 Excel export.

    Args:
        workbook: The workbook to add validation to
    """
    line_items_sheet = workbook["Line Items"]
    dictionary_sheet = workbook["Dictionary"]

    # Channel validation (existing)
    # channel_validation = DataValidation(
    #     type="list",
    #     formula1='"social,search,display,video,audio,tv,ooh,print,other"',
    #     allow_blank=True
    # )
    # line_items_sheet.add_data_validation(channel_validation)
    # Find channel column dynamically
    # for col in range(1, line_items_sheet.max_column + 1):
    #     if line_items_sheet.cell(1, col).value == "Channel":
    #         channel_validation.add(f'{get_column_letter(col)}2:{get_column_letter(col)}1000')
    #         break

    # KPI validation (existing)
    # kpi_validation = DataValidation(
    #     type="list",
    #     formula1='"CPM,CPC,CPA,CTR,CPV,CPI,ROAS,other"',
    #     allow_blank=True
    # )
    # line_items_sheet.add_data_validation(kpi_validation)
    # Find KPI column dynamically
    # for col in range(1, line_items_sheet.max_column + 1):
    #     if line_items_sheet.cell(1, col).value == "KPI":
    #         kpi_validation.add(f'{get_column_letter(col)}2:{get_column_letter(col)}1000')
    #         break

    # Note: Line Items location_type validation removed in v3.0
    # The location_type field at line item level is now legacy (kept for backwards compatibility)
    # Proper location targeting is defined in the Target Locations sheet

    # Target Locations: Location Type validation (expanded in v3.0)
    target_locations_sheet = workbook["Target Locations"]
    target_location_type_validation = DataValidation(
        type="list",
        formula1='"Country,State,DMA,County,Postcode,Radius,POI"',
        allow_blank=True
    )
    target_locations_sheet.add_data_validation(target_location_type_validation)
    # Find Location Type column dynamically
    for col in range(1, target_locations_sheet.max_column + 1):
        if target_locations_sheet.cell(1, col).value == "Location Type":
            target_location_type_validation.add(f'{get_column_letter(col)}2:{get_column_letter(col)}1000')
            break

    # Target Locations: Exclusion Type validation (expanded in v3.0)
    target_exclusion_type_validation = DataValidation(
        type="list",
        formula1='"Country,State,DMA,County,Postcode,Radius,POI"',
        allow_blank=True
    )
    target_locations_sheet.add_data_validation(target_exclusion_type_validation)
    # Find Exclusion Type column dynamically
    for col in range(1, target_locations_sheet.max_column + 1):
        if target_locations_sheet.cell(1, col).value == "Exclusion Type":
            target_exclusion_type_validation.add(f'{get_column_letter(col)}2:{get_column_letter(col)}1000')
            break

    # Target Audiences: Demo Gender validation (v3.0)
    target_audiences_sheet = workbook["Target Audiences"]
    demo_gender_validation = DataValidation(
        type="list",
        formula1='"Male,Female,Any"',
        allow_blank=True
    )
    target_audiences_sheet.add_data_validation(demo_gender_validation)
    # Find Demo Gender column dynamically
    for col in range(1, target_audiences_sheet.max_column + 1):
        if target_audiences_sheet.cell(1, col).value == "Demo Gender":
            demo_gender_validation.add(f'{get_column_letter(col)}2:{get_column_letter(col)}1000')
            break

    # Dictionary Status validation (updated for v3.0: 65 fields total, but only 40 have Status)
    # Structure: 5 meta + 5 campaign + 10 lineitem dims (rows 3-22) + 25 standard metrics (no status, rows 23-47) + 10 custom metrics (rows 48-57) + 10 custom costs (rows 58-67)
    # Apply validation to dimensions (rows 3-22) and custom fields (rows 48-67), excluding standard metrics which don't have Status
    status_validation = DataValidation(
        type="list",
        formula1='"enabled,disabled"',
        allow_blank=False
    )
    dictionary_sheet.add_data_validation(status_validation)
    status_validation.add('E3:E22')   # Custom dimensions (meta + campaign + lineitem)
    status_validation.add('E48:E67')  # Custom metrics and custom costs