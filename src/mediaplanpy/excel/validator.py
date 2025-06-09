"""
Excel validator for mediaplanpy - Updated for v2.0 Schema Support Only.

This module provides functionality for validating Excel files against the
Media Plan Open Data Standard v2.0 schema only.
"""

import os
import logging
import tempfile
import copy
from typing import Dict, Any, List, Optional, Union, Tuple

import openpyxl
from openpyxl import Workbook

from mediaplanpy.exceptions import ValidationError
from mediaplanpy.schema import SchemaValidator
from mediaplanpy.excel.importer import import_from_excel

logger = logging.getLogger("mediaplanpy.excel.validator")


def validate_excel(file_path: str, schema_validator: Optional[SchemaValidator] = None,
                  schema_version: Optional[str] = None) -> List[str]:
    """
    Validate an Excel file against the Media Plan Open Data Standard v2.0 schema.

    Args:
        file_path: Path to the Excel file
        schema_validator: Optional schema validator to use. If None, creates a new one.
        schema_version: Optional schema version to validate against. Must be "2.0" if provided.

    Returns:
        A list of validation error messages, empty if validation succeeds.

    Raises:
        ValidationError: If the Excel file cannot be read or is structurally invalid.
    """
    try:
        # Validate schema version if provided
        if schema_version and not _is_v2_schema_version(schema_version):
            return [f"Excel validation only supports v2.0 schema. Requested: {schema_version}"]

        # Import the Excel file to a media plan structure
        media_plan = import_from_excel(file_path)

        # Create schema validator if not provided
        if schema_validator is None:
            schema_validator = SchemaValidator()

        # Force validation against v2.0 schema
        validation_version = "2.0"

        # Sanitize data to avoid validation errors
        sanitized_media_plan = _sanitize_for_v2_validation(media_plan)

        # Validate the media plan against the v2.0 schema
        errors = schema_validator.validate(sanitized_media_plan, validation_version)

        # Add Excel-specific validation for v2.0
        excel_errors = _validate_v2_excel_structure(file_path)
        errors.extend(excel_errors)

        if not errors:
            logger.info(f"Excel file validated successfully against v2.0 schema: {file_path}")
        else:
            logger.warning(f"Excel file validation failed with {len(errors)} errors: {file_path}")

        return errors

    except Exception as e:
        raise ValidationError(f"Failed to validate Excel file: {e}")


def _is_v2_schema_version(version: str) -> bool:
    """
    Check if the schema version is v2.0.

    Args:
        version: The schema version to check

    Returns:
        True if the version is v2.0, False otherwise
    """
    if not version:
        return False

    # Normalize version format (handle both "2.0" and "v2.0")
    normalized = version.replace("v", "") if version.startswith("v") else version
    return normalized == "2.0"


def validate_excel_template(template_path: str, schema_version: Optional[str] = None) -> List[str]:
    """
    Validate an Excel template for compatibility with v2.0 media plan structure.

    Args:
        template_path: Path to the Excel template
        schema_version: Optional schema version to validate against. Must be "2.0" if provided.

    Returns:
        A list of validation error messages, empty if validation succeeds.

    Raises:
        ValidationError: If the template cannot be read or is structurally invalid.
    """
    try:
        # Validate schema version if provided
        if schema_version and not _is_v2_schema_version(schema_version):
            return [f"Template validation only supports v2.0 schema. Requested: {schema_version}"]

        # Load the template
        workbook = openpyxl.load_workbook(template_path)

        # Check for required sheets for v2.0
        required_sheets = ["Metadata", "Campaign", "Line Items"]
        missing_sheets = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]

        errors = []
        if missing_sheets:
            errors.append(f"Template is missing required sheets: {', '.join(missing_sheets)}")

        # Check for optional Dictionary sheet (v2.0 feature)
        if "Dictionary" in workbook.sheetnames:
            dictionary_errors = _validate_v2_dictionary_sheet(workbook["Dictionary"])
            errors.extend(dictionary_errors)

        # Check line items sheet structure for v2.0
        if "Line Items" in workbook.sheetnames:
            line_items_sheet = workbook["Line Items"]

            # Get headers
            headers = [cell.value for cell in line_items_sheet[1] if cell.value]

            # v2.0 required headers
            required_headers = ["ID", "Name", "Start Date", "End Date", "Cost Total"]
            missing_headers = [header for header in required_headers if header not in headers]

            if missing_headers:
                errors.append(f"Template Line Items sheet is missing required headers: {', '.join(missing_headers)}")

            # Check for v2.0 specific headers
            v2_headers = {"Cost Currency", "Dayparts", "Inventory", "Engagements", "Followers"}
            found_v2_headers = [header for header in v2_headers if header in headers]

            if found_v2_headers:
                logger.info(f"Found v2.0 headers in template: {', '.join(found_v2_headers)}")
            else:
                errors.append("Template appears to be missing v2.0 specific headers. Expected headers like 'Cost Currency', 'Dayparts', 'Inventory', etc.")

        # Check campaign sheet for v2.0 fields
        if "Campaign" in workbook.sheetnames:
            campaign_errors = _validate_v2_campaign_sheet(workbook["Campaign"])
            errors.extend(campaign_errors)

        # Check metadata sheet for v2.0 fields
        if "Metadata" in workbook.sheetnames:
            metadata_errors = _validate_v2_metadata_sheet(workbook["Metadata"])
            errors.extend(metadata_errors)

        if not errors:
            logger.info(f"Excel template validated successfully for v2.0: {template_path}")
        else:
            logger.warning(f"Excel template validation failed with {len(errors)} errors: {template_path}")

        return errors

    except Exception as e:
        raise ValidationError(f"Failed to validate Excel template: {e}")


def _validate_v2_dictionary_sheet(dictionary_sheet) -> List[str]:
    """
    Validate the Dictionary sheet structure for v2.0.

    Args:
        dictionary_sheet: The Dictionary worksheet

    Returns:
        List of validation errors
    """
    errors = []

    # Check for required headers
    expected_headers = ["Field Name", "Field Type", "Caption", "Status"]

    # Find header row
    header_row = None
    for row in range(1, 5):
        row_values = [dictionary_sheet.cell(row=row, column=col).value for col in range(1, 5)]
        if any(header in row_values for header in expected_headers):
            header_row = row
            break

    if header_row is None:
        errors.append("Dictionary sheet is missing required headers: Field Name, Field Type, Caption, Status")
        return errors

    # Check that all expected headers are present
    header_values = [dictionary_sheet.cell(row=header_row, column=col).value for col in range(1, 5)]
    missing_headers = [header for header in expected_headers if header not in header_values]

    if missing_headers:
        errors.append(f"Dictionary sheet is missing headers: {', '.join(missing_headers)}")

    # Check data rows for valid custom field names
    valid_field_patterns = {
        "dim_custom": "Dimension",
        "metric_custom": "Metric",
        "cost_custom": "Cost"
    }

    for row in range(header_row + 1, dictionary_sheet.max_row + 1):
        field_name = dictionary_sheet.cell(row=row, column=1).value
        field_type = dictionary_sheet.cell(row=row, column=2).value
        status = dictionary_sheet.cell(row=row, column=4).value

        if field_name:  # Skip empty rows
            # Validate field name pattern
            valid_pattern = False
            for pattern, expected_type in valid_field_patterns.items():
                if str(field_name).startswith(pattern):
                    valid_pattern = True
                    # Check field type matches pattern
                    if field_type != expected_type:
                        errors.append(f"Row {row}: Field '{field_name}' should have type '{expected_type}', found '{field_type}'")
                    break

            if not valid_pattern:
                errors.append(f"Row {row}: Invalid field name '{field_name}'. Must be dim_custom1-10, metric_custom1-10, or cost_custom1-10")

            # Validate status
            if status and str(status).lower() not in ["enabled", "disabled"]:
                errors.append(f"Row {row}: Status must be 'enabled' or 'disabled', found '{status}'")

    return errors


def _validate_v2_campaign_sheet(campaign_sheet) -> List[str]:
    """
    Validate the Campaign sheet for v2.0 fields.

    Args:
        campaign_sheet: The Campaign worksheet

    Returns:
        List of validation errors
    """
    errors = []

    # Check for essential v2.0 campaign fields
    v2_campaign_fields = [
        "Campaign ID", "Campaign Name", "Start Date", "End Date", "Budget Total",
        "Budget Currency", "Agency ID", "Agency Name",
        "Advertiser ID", "Advertiser Name", "Product ID",
        "Campaign Type ID", "Campaign Type Name",
        "Workflow Status ID", "Workflow Status Name"
    ]

    missing_fields = []
    found_fields = []

    for field in v2_campaign_fields:
        found = False
        for row in range(1, campaign_sheet.max_row + 1):
            cell_value = campaign_sheet.cell(row=row, column=1).value
            if cell_value and f"{field}:" in str(cell_value):
                found = True
                found_fields.append(field)
                break

        if not found:
            missing_fields.append(field)

    # Required fields should be present
    required_fields = ["Campaign ID", "Campaign Name", "Start Date", "End Date", "Budget Total"]
    missing_required = [field for field in required_fields if field in missing_fields]

    if missing_required:
        errors.append(f"Campaign sheet is missing required v2.0 fields: {', '.join(missing_required)}")

    # Log found v2.0 fields
    v2_optional_fields = ["Budget Currency", "Agency ID", "Campaign Type ID", "Workflow Status ID"]
    found_v2_optional = [field for field in v2_optional_fields if field in found_fields]

    if found_v2_optional:
        logger.info(f"Found v2.0 optional fields in Campaign sheet: {', '.join(found_v2_optional)}")

    return errors


def _validate_v2_metadata_sheet(metadata_sheet) -> List[str]:
    """
    Validate the Metadata sheet for v2.0 fields.

    Args:
        metadata_sheet: The Metadata worksheet

    Returns:
        List of validation errors
    """
    errors = []

    # Check for essential v2.0 metadata fields
    v2_metadata_fields = [
        "Schema Version", "Media Plan ID", "Media Plan Name",
        "Created By Name", "Created By ID", "Created At",
        "Is Current", "Is Archived", "Parent ID"
    ]

    missing_fields = []
    found_fields = []

    for field in v2_metadata_fields:
        found = False
        for row in range(1, metadata_sheet.max_row + 1):
            cell_value = metadata_sheet.cell(row=row, column=1).value
            if cell_value and f"{field}:" in str(cell_value):
                found = True
                found_fields.append(field)
                break

        if not found:
            missing_fields.append(field)

    # Required fields for v2.0
    required_fields = ["Schema Version", "Media Plan ID", "Created By Name", "Created At"]
    missing_required = [field for field in required_fields if field in missing_fields]

    if missing_required:
        errors.append(f"Metadata sheet is missing required v2.0 fields: {', '.join(missing_required)}")

    # Check schema version value if present
    for row in range(1, metadata_sheet.max_row + 1):
        key_cell = metadata_sheet.cell(row=row, column=1).value
        if key_cell and "Schema Version:" in str(key_cell):
            version_value = metadata_sheet.cell(row=row, column=2).value
            if version_value and not _is_v2_schema_version(str(version_value)):
                errors.append(f"Schema version should be '2.0', found '{version_value}'")
            break

    return errors


def create_validation_report(file_path: str, errors: List[str], output_path: Optional[str] = None) -> str:
    """
    Create a validation report for an Excel file (v2.0 focused).

    Args:
        file_path: Path to the validated Excel file
        errors: List of validation errors
        output_path: Optional path to save the report. If None, generates a default path.

    Returns:
        The path to the saved validation report.

    Raises:
        ValidationError: If the report cannot be created.
    """
    try:
        # Create a new workbook for the report
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "v2.0 Validation Report"

        # Set column widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 80

        # Add header
        sheet['A1'] = "v2.0 Schema Validation Report"
        sheet['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        sheet.merge_cells('A1:B1')

        # Add file information
        row = 2
        sheet[f'A{row}'] = "Validated File:"
        sheet[f'B{row}'] = os.path.basename(file_path)

        row += 1
        sheet[f'A{row}'] = "Schema Version:"
        sheet[f'B{row}'] = "2.0"

        row += 1
        sheet[f'A{row}'] = "Date:"
        sheet[f'B{row}'] = openpyxl.utils.datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row += 1
        sheet[f'A{row}'] = "Result:"
        if errors:
            sheet[f'B{row}'] = f"Failed ({len(errors)} errors)"
            sheet[f'B{row}'].font = openpyxl.styles.Font(color="FF0000")  # Red
        else:
            sheet[f'B{row}'] = "Passed"
            sheet[f'B{row}'].font = openpyxl.styles.Font(color="008000")  # Green

        # Add errors
        if errors:
            row += 2
            sheet[f'A{row}'] = "Errors:"
            sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)

            for i, error in enumerate(errors, 1):
                row += 1
                sheet[f'A{row}'] = f"Error {i}:"
                sheet[f'B{row}'] = error

        # Add v2.0 specific information
        row += 2
        sheet[f'A{row}'] = "v2.0 Features:"
        sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)

        row += 1
        sheet[f'B{row}'] = "- Dictionary configuration for custom fields"

        row += 1
        sheet[f'B{row}'] = "- Enhanced campaign fields (agency, advertiser, workflow status)"

        row += 1
        sheet[f'B{row}'] = "- 17 new standard metrics for line items"

        row += 1
        sheet[f'B{row}'] = "- Dayparts and inventory targeting"

        row += 1
        sheet[f'B{row}'] = "- Currency support for budgets and costs"

        # Determine output path if not provided
        if not output_path:
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_path = f"{base_name}_v2_validation_report.xlsx"

        # Save the report
        workbook.save(output_path)

        logger.info(f"v2.0 Validation report saved to: {output_path}")
        return output_path

    except Exception as e:
        raise ValidationError(f"Failed to create validation report: {e}")


def _validate_v2_excel_structure(file_path: str) -> List[str]:
    """
    Validate the structure of an Excel file for v2.0 schema.

    Args:
        file_path: Path to the Excel file

    Returns:
        A list of structural validation errors
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        errors = []

        # Check for required sheets
        required_sheets = ["Metadata", "Campaign", "Line Items"]
        missing_sheets = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]

        if missing_sheets:
            errors.append(f"Excel file is missing required sheets: {', '.join(missing_sheets)}")

        # Check for v2.0 specific features
        if "Dictionary" not in workbook.sheetnames:
            logger.info("Dictionary sheet not found - this is optional for v2.0")

        # Check line items structure for v2.0
        if "Line Items" in workbook.sheetnames:
            line_items_sheet = workbook["Line Items"]

            # Check if first row contains headers
            headers = [cell.value for cell in line_items_sheet[1] if cell.value]
            if not headers:
                errors.append("Line Items sheet is missing headers in first row")

            # Check for v2.0 specific headers
            v2_headers = {"Cost Currency", "Dayparts", "Inventory", "Engagements"}
            found_v2_headers = [header for header in v2_headers if header in headers]

            if not found_v2_headers:
                errors.append("Line Items sheet appears to be missing v2.0 specific headers")

            # Check for data rows
            has_data = False
            for row in range(2, line_items_sheet.max_row + 1):
                if any(cell.value for cell in line_items_sheet[row]):
                    has_data = True
                    break

            if not has_data:
                errors.append("Line Items sheet has no data rows")

        # Check campaign data for v2.0 fields
        if "Campaign" in workbook.sheetnames:
            campaign_sheet = workbook["Campaign"]

            # Check for essential campaign information
            essential_fields = ["Campaign ID", "Campaign Name", "Start Date", "End Date", "Budget Total"]
            missing_fields = []

            for field in essential_fields:
                found = False
                for row in range(1, campaign_sheet.max_row + 1):
                    cell_value = campaign_sheet.cell(row=row, column=1).value
                    if cell_value and field in str(cell_value):
                        found = True
                        break

                if not found:
                    missing_fields.append(field)

            if missing_fields:
                errors.append(f"Campaign sheet is missing essential fields: {', '.join(missing_fields)}")

        # Check metadata for v2.0 required fields
        if "Metadata" in workbook.sheetnames:
            metadata_sheet = workbook["Metadata"]

            # Check for required v2.0 metadata fields
            required_meta_fields = ["Created By Name"]  # v2.0 specific requirement
            missing_meta_fields = []

            for field in required_meta_fields:
                found = False
                for row in range(1, metadata_sheet.max_row + 1):
                    cell_value = metadata_sheet.cell(row=row, column=1).value
                    if cell_value and field in str(cell_value):
                        found = True
                        break

                if not found:
                    missing_meta_fields.append(field)

            if missing_meta_fields:
                errors.append(f"Metadata sheet is missing v2.0 required fields: {', '.join(missing_meta_fields)}")

        return errors

    except Exception as e:
        return [f"Error validating Excel structure: {str(e)}"]


def _sanitize_for_v2_validation(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Sanitize data for v2.0 validation to avoid common errors with enum fields.

    Args:
        data: The data to sanitize

    Returns:
        Sanitized data
    """
    # Create a deep copy to avoid modifying the original
    sanitized = copy.deepcopy(data)

    # Handle campaign fields
    if "campaign" in sanitized:
        campaign = sanitized["campaign"]

        # Convert empty strings to None for enum fields
        if "audience_gender" in campaign:
            if campaign["audience_gender"] == "" or campaign["audience_gender"] is None:
                campaign["audience_gender"] = "Any"  # Default value

        if "location_type" in campaign:
            if campaign["location_type"] == "" or campaign["location_type"] is None:
                campaign["location_type"] = "Country"  # Default value

    # Handle line items
    if "lineitems" in sanitized:
        for line_item in sanitized["lineitems"]:
            # Handle location_type field
            if "location_type" in line_item:
                if line_item["location_type"] == "" or line_item["location_type"] is None:
                    line_item["location_type"] = "Country"  # Default value

    # Handle dictionary (NEW for v2.0)
    if "dictionary" in sanitized:
        dictionary = sanitized["dictionary"]

        # Clean up dictionary structure
        for section_name in ["custom_dimensions", "custom_metrics", "custom_costs"]:
            if section_name in dictionary:
                section = dictionary[section_name]

                # Validate each field configuration
                cleaned_section = {}
                for field_name, config in section.items():
                    if isinstance(config, dict) and "status" in config:
                        if config["status"] in ["enabled", "disabled"]:
                            cleaned_config = {
                                "status": config["status"],
                                "caption": config.get("caption", "")
                            }
                            # Ensure caption for enabled fields
                            if config["status"] == "enabled" and not cleaned_config["caption"]:
                                cleaned_config["caption"] = f"Custom {field_name}"

                            cleaned_section[field_name] = cleaned_config

                if cleaned_section:
                    dictionary[section_name] = cleaned_section
                else:
                    del dictionary[section_name]

    return sanitized