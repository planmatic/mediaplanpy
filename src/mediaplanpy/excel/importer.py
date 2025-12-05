"""
Excel importer for mediaplanpy - Updated for v3.0 Schema Support Only.

This module provides functionality for importing media plans from Excel format,
supporting only v3.0 schema with target audiences/locations and enhanced dictionary.
"""

import os
import logging
import uuid
from datetime import datetime, date
from typing import Dict, Any, List, Optional, Union, Tuple
from decimal import Decimal
import copy

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from mediaplanpy.exceptions import StorageError, ValidationError

logger = logging.getLogger("mediaplanpy.excel.importer")


def _parse_json_field(value: Any, field_name: str) -> Dict[str, Any]:
    """
    Parse a JSON field from Excel, validating and returning {} if empty or invalid.

    Args:
        value: The cell value (could be string, dict, None, etc.)
        field_name: Name of the field for logging

    Returns:
        Parsed JSON as dictionary, or {} if empty/invalid
    """
    import json

    # If empty, return empty dict
    if not value or (isinstance(value, str) and not value.strip()):
        return {}

    # If already a dict, return it
    if isinstance(value, dict):
        return value

    # Try to parse as JSON string
    try:
        parsed = json.loads(str(value))
        if isinstance(parsed, dict):
            return parsed
        else:
            logger.warning(f"Field '{field_name}' contains non-object JSON, returning empty dict")
            return {}
    except (json.JSONDecodeError, ValueError) as e:
        logger.warning(f"Field '{field_name}' contains invalid JSON: {e}, returning empty dict")
        return {}


def import_from_excel(file_path: str, **kwargs) -> Dict[str, Any]:
    """
    Import a media plan from an Excel file using v3.0 schema with enhanced data validation.

    Args:
        file_path: Path to the Excel file
        **kwargs: Additional import options

    Returns:
        The imported media plan data in v3.0 schema format

    Raises:
        StorageError: If the import fails
        ValidationError: If the Excel file contains invalid data or wrong schema version
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Detect and validate schema version
        schema_version = _detect_schema_version(workbook)
        if not _is_v3_schema_version(schema_version):
            raise ValidationError(
                f"Excel import only supports v3.0 schema. Found: {schema_version}. "
                f"Please use a v3.0 compatible Excel file or upgrade your v2.0 file first."
            )

        # Import v3.0 media plan
        media_plan = _import_v3_media_plan(workbook)

        # Perform data integrity validations
        _validate_import_data_integrity(media_plan)

        # Sanitize the data to avoid validation issues
        sanitized_media_plan = _sanitize_v3_media_plan_data(media_plan)

        logger.info(f"Media plan imported from Excel (v3.0): {file_path}")
        return sanitized_media_plan

    except ValidationError:
        # Re-raise validation errors as-is to preserve detailed messages
        raise
    except Exception as e:
        raise StorageError(f"Failed to import media plan from Excel: {e}")


def _validate_import_data_integrity(media_plan: Dict[str, Any]) -> None:
    """
    Validate data integrity of imported media plan data.

    Args:
        media_plan: The imported media plan data to validate

    Raises:
        ValidationError: If validation fails with detailed error messages
    """
    validation_errors = []

    # Validate Line Item ID uniqueness
    lineitem_errors = _validate_lineitem_id_uniqueness(media_plan.get("lineitems", []))
    validation_errors.extend(lineitem_errors)

    # Validate Dictionary caption uniqueness
    dictionary_errors = _validate_dictionary_caption_uniqueness(media_plan.get("dictionary", {}))
    validation_errors.extend(dictionary_errors)

    # If any validation errors found, raise ValidationError
    if validation_errors:
        error_message = "Data integrity validation failed:\n\n" + "\n".join(validation_errors)
        raise ValidationError(error_message)


def _validate_lineitem_id_uniqueness(lineitems: List[Dict[str, Any]]) -> List[str]:
    """
    Validate that all non-empty Line Item IDs are unique.

    Args:
        lineitems: List of line item dictionaries

    Returns:
        List of validation error messages (empty if no errors)
    """
    errors = []

    if not lineitems:
        return errors

    # Collect non-empty IDs with their row numbers (Excel row = list index + 2 for header)
    id_to_rows = {}

    for idx, lineitem in enumerate(lineitems):
        line_id = lineitem.get("id")
        excel_row = idx + 2  # +2 because Excel has header row and is 1-indexed

        # Only check non-empty IDs (empty/None IDs will get auto-generated)
        if line_id and str(line_id).strip():
            clean_id = str(line_id).strip()

            if clean_id not in id_to_rows:
                id_to_rows[clean_id] = []
            id_to_rows[clean_id].append(excel_row)

    # Find duplicates
    duplicates = {id_val: rows for id_val, rows in id_to_rows.items() if len(rows) > 1}

    if duplicates:
        errors.append("❌ Line Item ID Uniqueness Validation Failed:")
        errors.append("   The following Line Item IDs appear multiple times:")

        for duplicate_id, rows in duplicates.items():
            row_list = ", ".join(f"row {row}" for row in sorted(rows))
            errors.append(f"   • ID '{duplicate_id}' found in: {row_list}")

        errors.append("")
        errors.append("   💡 Solution: Ensure each Line Item has a unique ID, or leave ID blank for auto-generation.")

    return errors


def _validate_dictionary_caption_uniqueness(dictionary: Dict[str, Any]) -> List[str]:
    """
    Validate that custom field captions are unique across ALL custom field types.

    Args:
        dictionary: Dictionary configuration data

    Returns:
        List of validation error messages (empty if no errors)
    """
    errors = []

    if not dictionary:
        return errors

    # Define section information for clear error messages
    sections = {
        "custom_dimensions": {
            "name": "Custom Dimensions",
            "field_prefix": "dim_custom"
        },
        "custom_metrics": {
            "name": "Custom Metrics",
            "field_prefix": "metric_custom"
        },
        "custom_costs": {
            "name": "Custom Costs",
            "field_prefix": "cost_custom"
        }
    }

    # Collect ALL captions from ALL sections (global uniqueness check)
    caption_to_fields = {}

    for section_key, section_info in sections.items():
        section_data = dictionary.get(section_key, {})

        if not section_data:
            continue

        for field_name, config in section_data.items():
            if not isinstance(config, dict):
                continue

            status = config.get("status", "").lower()
            caption = config.get("caption", "").strip()

            # Only check enabled fields with non-empty captions
            if status == "enabled" and caption:
                if caption not in caption_to_fields:
                    caption_to_fields[caption] = []

                # Store field name with section information for better error reporting
                field_info = {
                    "field_name": field_name,
                    "section": section_info["name"]
                }
                caption_to_fields[caption].append(field_info)

    # Find captions used by multiple fields (across ALL sections)
    duplicates = {caption: fields for caption, fields in caption_to_fields.items() if len(fields) > 1}

    if duplicates:
        errors.append("❌ Dictionary Caption Uniqueness Validation Failed:")
        errors.append("   The following captions are used by multiple custom fields:")
        errors.append("")

        for duplicate_caption, field_infos in duplicates.items():
            errors.append(f"   Caption '{duplicate_caption}' is used by:")

            # Group by section for cleaner display
            by_section = {}
            for field_info in field_infos:
                section = field_info["section"]
                if section not in by_section:
                    by_section[section] = []
                by_section[section].append(field_info["field_name"])

            # Display grouped by section
            for section, field_names in by_section.items():
                field_list = ", ".join(sorted(field_names))
                errors.append(f"      • {section}: {field_list}")

            errors.append("")  # Add blank line between different caption conflicts

        errors.append("   💡 Solution: Each custom field must have a unique caption across ALL custom field types.")
        errors.append("      Captions must be unique globally - not just within dimensions, metrics, or costs.")

    return errors


def update_from_excel(media_plan: Dict[str, Any], file_path: str, **kwargs) -> Dict[str, Any]:
    """
    Update a media plan from an Excel file using v3.0 schema with enhanced data validation.

    Args:
        media_plan: The existing media plan data to update
        file_path: Path to the Excel file
        **kwargs: Additional import options

    Returns:
        The updated media plan data in v3.0 schema format

    Raises:
        StorageError: If the update fails
        ValidationError: If the Excel file contains invalid data
    """
    try:
        # Import the Excel file (this will include our new validations)
        updated_plan = import_from_excel(file_path, **kwargs)

        # Keep original meta data, but update relevant fields
        result = media_plan.copy()

        # Preserve original meta ID and timestamps, but update other meta fields
        original_meta = result.get("meta", {})
        updated_meta = updated_plan.get("meta", {})

        # Keep original ID and timestamps
        updated_meta["id"] = original_meta.get("id", updated_meta.get("id"))
        updated_meta["created_at"] = original_meta.get("created_at", updated_meta.get("created_at"))

        # Update meta with new values from Excel
        result["meta"] = updated_meta

        # Update campaign and line items
        result["campaign"] = updated_plan["campaign"]
        result["lineitems"] = updated_plan["lineitems"]

        # Update dictionary if present
        if "dictionary" in updated_plan:
            result["dictionary"] = updated_plan["dictionary"]

        logger.info(f"Media plan updated from Excel (v3.0): {file_path}")
        return result

    except ValidationError:
        # Re-raise validation errors as-is to preserve detailed messages
        raise
    except Exception as e:
        raise StorageError(f"Failed to update media plan from Excel: {e}")


def _is_v3_schema_version(version: str) -> bool:
    """
    Check if the schema version is v3.0.

    Args:
        version: The schema version to check

    Returns:
        True if the version is v3.0, False otherwise
    """
    if not version:
        return False

    # Normalize version format (handle both "3.0" and "v3.0")
    normalized = version.replace("v", "") if version.startswith("v") else version
    return normalized == "3.0"


def _detect_schema_version(workbook: Workbook) -> str:
    """
    Detect the schema version from the Metadata sheet in an Excel workbook.

    Args:
        workbook: The workbook to analyze

    Returns:
        The detected schema version, or "unknown" if not found
    """
    # Check if Metadata sheet exists
    if "Metadata" not in workbook.sheetnames:
        logger.warning("Metadata sheet not found in Excel workbook")
        return "unknown"

    metadata_sheet = workbook["Metadata"]

    # Look for Schema Version in the Metadata sheet (column A = key, column B = value)
    for row in range(1, 20):  # Check first 20 rows
        key_cell = metadata_sheet.cell(row=row, column=1).value
        if key_cell == "Schema Version:":
            version = metadata_sheet.cell(row=row, column=2).value
            if version:
                # Normalize version format (remove 'v' prefix if present)
                normalized = version.replace("v", "") if isinstance(version, str) and version.startswith("v") else str(version)
                logger.info(f"Detected schema version: {normalized}")
                return normalized

    # Schema version not found in Metadata sheet
    logger.warning("Schema Version field not found in Metadata sheet")
    return "unknown"


def _import_v3_media_plan(workbook: Workbook) -> Dict[str, Any]:
    """
    Import a v3.0 media plan from a workbook.

    Args:
        workbook: The workbook to import from

    Returns:
        The imported media plan data in v3.0 format
    """
    media_plan = {
        "meta": {},
        "campaign": {},
        "lineitems": [],
        "dictionary": {}
    }

    # Import metadata with v3.0 fields (dim_custom1-5, custom_properties)
    if "Metadata" in workbook.sheetnames:
        metadata_sheet = workbook["Metadata"]
        meta = _import_v3_metadata(metadata_sheet)
        media_plan["meta"] = meta

    # Import campaign with v3.0 fields (target_audiences, target_locations arrays, KPIs, etc.)
    if "Campaign" in workbook.sheetnames:
        campaign_sheet = workbook["Campaign"]
        campaign = _import_v3_campaign(campaign_sheet)

        # Import Target Audiences (optional - if sheet exists)
        if "Target Audiences" in workbook.sheetnames:
            target_audiences_sheet = workbook["Target Audiences"]
            campaign["target_audiences"] = _import_v3_target_audiences(target_audiences_sheet)
        else:
            campaign["target_audiences"] = []

        # Import Target Locations (optional - if sheet exists)
        if "Target Locations" in workbook.sheetnames:
            target_locations_sheet = workbook["Target Locations"]
            campaign["target_locations"] = _import_v3_target_locations(target_locations_sheet)
        else:
            campaign["target_locations"] = []

        media_plan["campaign"] = campaign

    # Import line items with v3.0 fields (buy fields, new metrics, JSON fields)
    if "Line Items" in workbook.sheetnames:
        line_items_sheet = workbook["Line Items"]
        lineitems = _import_v3_lineitems(line_items_sheet)
        media_plan["lineitems"] = lineitems

    # Import dictionary configuration (v3.0: 6 sections with formula support)
    if "Dictionary" in workbook.sheetnames:
        dictionary_sheet = workbook["Dictionary"]
        dictionary = _import_v3_dictionary(dictionary_sheet)
        if dictionary:  # Only include if not empty
            media_plan["dictionary"] = dictionary

    # Ensure required meta fields with v3.0 format
    _ensure_v3_meta_fields(media_plan["meta"])

    # Ensure required campaign fields with v3.0 format
    _ensure_v3_campaign_fields(media_plan["campaign"])

    return media_plan


def _import_v3_metadata(metadata_sheet) -> Dict[str, Any]:
    """
    Import metadata section with v3.0 fields.

    Args:
        metadata_sheet: The metadata worksheet

    Returns:
        Dictionary containing metadata
    """
    meta = {}

    for row in range(1, 30):  # Check first 30 rows (more fields in v3.0)
        key_cell = metadata_sheet.cell(row=row, column=1).value
        value_cell = metadata_sheet.cell(row=row, column=2).value

        if key_cell == "Schema Version:":
            meta["schema_version"] = str(value_cell) if value_cell else "3.0"
        elif key_cell == "Media Plan ID:":
            meta["id"] = value_cell or f"mediaplan_{uuid.uuid4().hex[:8]}"
        elif key_cell == "Media Plan Name:":
            meta["name"] = value_cell or ""
        elif key_cell == "Created By Name:":
            meta["created_by_name"] = value_cell or ""
        elif key_cell == "Created By ID:":
            meta["created_by_id"] = value_cell or None
        elif key_cell == "Created At:":
            meta["created_at"] = value_cell or datetime.now().isoformat()
        elif key_cell == "Is Current:":
            if value_cell is not None:
                meta["is_current"] = str(value_cell).lower() in ['true', 'yes', '1']
        elif key_cell == "Is Archived:":
            if value_cell is not None:
                meta["is_archived"] = str(value_cell).lower() in ['true', 'yes', '1']
        elif key_cell == "Parent ID:":
            meta["parent_id"] = value_cell or None
        elif key_cell == "Comments:":
            meta["comments"] = value_cell or ""
        # NEW v3.0 fields: dim_custom1-5
        elif key_cell == "Dim Custom 1:":
            if value_cell and str(value_cell).strip():
                meta["dim_custom1"] = str(value_cell).strip()
        elif key_cell == "Dim Custom 2:":
            if value_cell and str(value_cell).strip():
                meta["dim_custom2"] = str(value_cell).strip()
        elif key_cell == "Dim Custom 3:":
            if value_cell and str(value_cell).strip():
                meta["dim_custom3"] = str(value_cell).strip()
        elif key_cell == "Dim Custom 4:":
            if value_cell and str(value_cell).strip():
                meta["dim_custom4"] = str(value_cell).strip()
        elif key_cell == "Dim Custom 5:":
            if value_cell and str(value_cell).strip():
                meta["dim_custom5"] = str(value_cell).strip()
        # NEW v3.0 field: custom_properties (JSON)
        elif key_cell == "Custom Properties (JSON):":
            if value_cell:
                meta["custom_properties"] = _parse_json_field(value_cell, "meta.custom_properties")

    return meta


def _import_v3_campaign(campaign_sheet) -> Dict[str, Any]:
    """
    Import campaign section with v3.0 fields.

    NOTE: target_audiences and target_locations are imported from separate sheets
    by the orchestrator function, not from this Campaign sheet.

    Args:
        campaign_sheet: The campaign worksheet

    Returns:
        Dictionary containing campaign data
    """
    campaign = {}

    for row in range(1, 70):  # Check more rows for v3.0 fields (more KPIs)
        key_cell = campaign_sheet.cell(row=row, column=1).value
        value_cell = campaign_sheet.cell(row=row, column=2).value

        if key_cell == "Campaign ID:":
            campaign["id"] = value_cell or f"campaign_{uuid.uuid4().hex[:8]}"
        elif key_cell == "Campaign Name:":
            campaign["name"] = value_cell or ""
        elif key_cell == "Objective:":
            campaign["objective"] = value_cell or ""
        elif key_cell == "Start Date:":
            if isinstance(value_cell, date):
                campaign["start_date"] = value_cell.isoformat()
            else:
                campaign["start_date"] = str(value_cell) if value_cell else ""
        elif key_cell == "End Date:":
            if isinstance(value_cell, date):
                campaign["end_date"] = value_cell.isoformat()
            else:
                campaign["end_date"] = str(value_cell) if value_cell else ""
        elif key_cell == "Budget Total:":
            campaign["budget_total"] = float(value_cell) if value_cell else 0

        # NEW v2.0 budget field
        elif key_cell == "Budget Currency:":
            if value_cell and str(value_cell).strip():
                campaign["budget_currency"] = str(value_cell).strip()

        # NEW v2.0 agency fields
        elif key_cell == "Agency ID:":
            if value_cell and str(value_cell).strip():
                campaign["agency_id"] = str(value_cell).strip()
        elif key_cell == "Agency Name:":
            if value_cell and str(value_cell).strip():
                campaign["agency_name"] = str(value_cell).strip()

        # NEW v2.0 advertiser fields
        elif key_cell == "Advertiser ID:":
            if value_cell and str(value_cell).strip():
                campaign["advertiser_id"] = str(value_cell).strip()
        elif key_cell == "Advertiser Name:":
            if value_cell and str(value_cell).strip():
                campaign["advertiser_name"] = str(value_cell).strip()

        # Product fields (existing + new v2.0 product_id)
        elif key_cell == "Product ID:":
            if value_cell and str(value_cell).strip():
                campaign["product_id"] = str(value_cell).strip()
        elif key_cell == "Product Name:":
            if value_cell and str(value_cell).strip():
                campaign["product_name"] = str(value_cell).strip()
        elif key_cell == "Product Description:":
            if value_cell and str(value_cell).strip():
                campaign["product_description"] = str(value_cell).strip()

        # NEW v2.0 campaign type fields
        elif key_cell == "Campaign Type ID:":
            if value_cell and str(value_cell).strip():
                campaign["campaign_type_id"] = str(value_cell).strip()
        elif key_cell == "Campaign Type Name:":
            if value_cell and str(value_cell).strip():
                campaign["campaign_type_name"] = str(value_cell).strip()

        # Workflow status fields
        elif key_cell == "Workflow Status ID:":
            if value_cell and str(value_cell).strip():
                campaign["workflow_status_id"] = str(value_cell).strip()
        elif key_cell == "Workflow Status Name:":
            if value_cell and str(value_cell).strip():
                campaign["workflow_status_name"] = str(value_cell).strip()

        # NEW v3.0 KPI fields (5 pairs)
        elif key_cell == "KPI Name 1:":
            if value_cell and str(value_cell).strip():
                campaign["kpi_name1"] = str(value_cell).strip()
        elif key_cell == "KPI Value 1:":
            if value_cell:
                try:
                    campaign["kpi_value1"] = float(value_cell)
                except (ValueError, TypeError):
                    pass
        elif key_cell == "KPI Name 2:":
            if value_cell and str(value_cell).strip():
                campaign["kpi_name2"] = str(value_cell).strip()
        elif key_cell == "KPI Value 2:":
            if value_cell:
                try:
                    campaign["kpi_value2"] = float(value_cell)
                except (ValueError, TypeError):
                    pass
        elif key_cell == "KPI Name 3:":
            if value_cell and str(value_cell).strip():
                campaign["kpi_name3"] = str(value_cell).strip()
        elif key_cell == "KPI Value 3:":
            if value_cell:
                try:
                    campaign["kpi_value3"] = float(value_cell)
                except (ValueError, TypeError):
                    pass
        elif key_cell == "KPI Name 4:":
            if value_cell and str(value_cell).strip():
                campaign["kpi_name4"] = str(value_cell).strip()
        elif key_cell == "KPI Value 4:":
            if value_cell:
                try:
                    campaign["kpi_value4"] = float(value_cell)
                except (ValueError, TypeError):
                    pass
        elif key_cell == "KPI Name 5:":
            if value_cell and str(value_cell).strip():
                campaign["kpi_name5"] = str(value_cell).strip()
        elif key_cell == "KPI Value 5:":
            if value_cell:
                try:
                    campaign["kpi_value5"] = float(value_cell)
                except (ValueError, TypeError):
                    pass

        # NEW v3.0 custom dimension fields (dim_custom1-5)
        elif key_cell == "Dim Custom 1:":
            if value_cell and str(value_cell).strip():
                campaign["dim_custom1"] = str(value_cell).strip()
        elif key_cell == "Dim Custom 2:":
            if value_cell and str(value_cell).strip():
                campaign["dim_custom2"] = str(value_cell).strip()
        elif key_cell == "Dim Custom 3:":
            if value_cell and str(value_cell).strip():
                campaign["dim_custom3"] = str(value_cell).strip()
        elif key_cell == "Dim Custom 4:":
            if value_cell and str(value_cell).strip():
                campaign["dim_custom4"] = str(value_cell).strip()
        elif key_cell == "Dim Custom 5:":
            if value_cell and str(value_cell).strip():
                campaign["dim_custom5"] = str(value_cell).strip()

        # NEW v3.0 custom_properties (JSON)
        elif key_cell == "Custom Properties (JSON):":
            if value_cell:
                campaign["custom_properties"] = _parse_json_field(value_cell, "campaign.custom_properties")

    return campaign


def _import_v3_target_audiences(target_audiences_sheet) -> List[Dict[str, Any]]:
    """
    Import Target Audiences from v3.0 sheet (columnar format).

    Args:
        target_audiences_sheet: The Target Audiences worksheet

    Returns:
        List of target audience dictionaries
    """
    # Get headers from row 1
    headers = [cell.value for cell in target_audiences_sheet[1] if cell.value]

    # Field mapping for Target Audiences (13 columns)
    field_mapping = {
        "Name": "name",
        "Description": "description",
        "Demo Age Start": "demo_age_start",
        "Demo Age End": "demo_age_end",
        "Demo Gender": "demo_gender",
        "Demo Attributes": "demo_attributes",
        "Interest Attributes": "interest_attributes",
        "Intent Attributes": "intent_attributes",
        "Purchase Attributes": "purchase_attributes",
        "Content Attributes": "content_attributes",
        "Exclusion List": "exclusion_list",
        "Extension Approach": "extension_approach",
        "Population Size": "population_size"
    }

    # Map column indices to field names
    header_to_field = {}
    for col_idx, header in enumerate(headers, 1):
        if header in field_mapping:
            header_to_field[col_idx] = field_mapping[header]

    # Process rows (skip header row)
    audiences = []
    for row in range(2, target_audiences_sheet.max_row + 1):
        # Skip empty rows
        if all(cell.value is None for cell in target_audiences_sheet[row]):
            continue

        audience = {}
        for col_idx, field_name in header_to_field.items():
            cell_value = target_audiences_sheet.cell(row=row, column=col_idx).value

            if cell_value is not None:
                # Handle different field types
                if field_name in ["demo_age_start", "demo_age_end", "population_size"]:
                    try:
                        audience[field_name] = int(cell_value)
                    except (ValueError, TypeError):
                        pass
                elif field_name == "demo_gender":
                    if str(cell_value).strip():
                        audience[field_name] = str(cell_value).strip()
                else:
                    if str(cell_value).strip():
                        audience[field_name] = str(cell_value).strip()

        # Only add if has at least a name
        if audience.get("name"):
            audiences.append(audience)

    return audiences


def _import_v3_target_locations(target_locations_sheet) -> List[Dict[str, Any]]:
    """
    Import Target Locations from v3.0 sheet (columnar format).

    Args:
        target_locations_sheet: The Target Locations worksheet

    Returns:
        List of target location dictionaries
    """
    # Get headers from row 1
    headers = [cell.value for cell in target_locations_sheet[1] if cell.value]

    # Field mapping for Target Locations (7 columns)
    field_mapping = {
        "Name": "name",
        "Description": "description",
        "Location Type": "location_type",
        "Location List (JSON)": "location_list",
        "Exclusion Type": "exclusion_type",
        "Exclusion List (JSON)": "exclusion_list",
        "Population Percent": "population_percent"
    }

    # Map column indices to field names
    header_to_field = {}
    for col_idx, header in enumerate(headers, 1):
        if header in field_mapping:
            header_to_field[col_idx] = field_mapping[header]

    # Process rows (skip header row)
    locations = []
    for row in range(2, target_locations_sheet.max_row + 1):
        # Skip empty rows
        if all(cell.value is None for cell in target_locations_sheet[row]):
            continue

        location = {}
        for col_idx, field_name in header_to_field.items():
            cell_value = target_locations_sheet.cell(row=row, column=col_idx).value

            if cell_value is not None:
                # Handle different field types
                if field_name == "population_percent":
                    try:
                        location[field_name] = float(cell_value)
                    except (ValueError, TypeError):
                        pass
                elif field_name in ["location_list", "exclusion_list"]:
                    # Parse JSON array
                    import json
                    try:
                        if isinstance(cell_value, str) and cell_value.strip():
                            parsed = json.loads(cell_value)
                            if isinstance(parsed, list):
                                location[field_name] = parsed
                    except (json.JSONDecodeError, ValueError):
                        logger.warning(f"Invalid JSON in {field_name} field at row {row}, skipping")
                else:
                    if str(cell_value).strip():
                        location[field_name] = str(cell_value).strip()

        # Only add if has at least a name
        if location.get("name"):
            locations.append(location)

    return locations


def _import_v3_lineitems(line_items_sheet) -> List[Dict[str, Any]]:
    """
    Import line items section with v3.0 fields, including calculated columns for zero budget reconstruction.

    Args:
        line_items_sheet: The line items worksheet

    Returns:
        List of line item dictionaries
    """
    # Get headers
    headers = [cell.value for cell in line_items_sheet[1] if cell.value]

    # Create comprehensive field mapping for v3.0 - ONLY original schema fields
    field_mapping = {
        # Required fields
        "ID": "id",
        "Name": "name",
        "Start Date": "start_date",
        "End Date": "end_date",
        "Cost Total": "cost_total",

        # Channel-related fields
        "Channel": "channel",
        "Channel Custom": "channel_custom",
        "Vehicle": "vehicle",
        "Vehicle Custom": "vehicle_custom",
        "Partner": "partner",
        "Partner Custom": "partner_custom",
        "Media Product": "media_product",
        "Media Product Custom": "media_product_custom",

        # Location fields
        "Location Type": "location_type",
        "Location Name": "location_name",

        # Audience and format fields
        "Target Audience": "target_audience",
        "Ad Format": "adformat",
        "Ad Format Custom": "adformat_custom",

        # KPI fields
        "KPI": "kpi",
        "KPI Custom": "kpi_custom",
        "KPI Value": "kpi_value",  # NEW in v3.0

        # Dayparts and inventory fields
        "Dayparts": "dayparts",
        "Dayparts Custom": "dayparts_custom",
        "Inventory": "inventory",
        "Inventory Custom": "inventory_custom",

        # NEW v3.0: Buy type fields
        "Buy Type": "buy_type",
        "Buy Commitment": "buy_commitment",

        # Cost fields (including cost_currency)
        "Cost Currency": "cost_currency",
        "Cost Currency Exchange Rate": "cost_currency_exchange_rate",  # NEW in v3.0
        "Cost Media": "cost_media",
        "Cost Buying": "cost_buying",
        "Cost Platform": "cost_platform",
        "Cost Data": "cost_data",
        "Cost Creative": "cost_creative",
        "Cost Minimum": "cost_minimum",  # NEW in v3.0
        "Cost Maximum": "cost_maximum",  # NEW in v3.0

        # Metric fields - ALL v3.0 standard metrics
        "Impressions": "metric_impressions",
        "Clicks": "metric_clicks",
        "Views": "metric_views",
        "View Starts": "metric_view_starts",  # NEW in v3.0
        "View Completions": "metric_view_completions",  # NEW in v3.0
        "Reach": "metric_reach",  # NEW in v3.0
        "Units": "metric_units",  # NEW in v3.0
        "Impression Share": "metric_impression_share",  # NEW in v3.0
        "Engagements": "metric_engagements",
        "Followers": "metric_followers",
        "Visits": "metric_visits",
        "Leads": "metric_leads",
        "Sales": "metric_sales",
        "Add to Cart": "metric_add_to_cart",
        "App Install": "metric_app_install",
        "Application Start": "metric_application_start",
        "Application Complete": "metric_application_complete",
        "Contact Us": "metric_contact_us",
        "Download": "metric_download",
        "Signup": "metric_signup",
        "Page Views": "metric_page_views",  # NEW in v3.0
        "Likes": "metric_likes",  # NEW in v3.0
        "Shares": "metric_shares",  # NEW in v3.0
        "Comments": "metric_comments",  # NEW in v3.0
        "Conversions": "metric_conversions",  # NEW in v3.0
        "Max Daily Spend": "metric_max_daily_spend",
        "Max Daily Impressions": "metric_max_daily_impressions",
        "Audience Size": "metric_audience_size",

        # NEW v3.0: Aggregate fields
        "Is Aggregate": "is_aggregate",
        "Aggregation Level": "aggregation_level",

        # NEW v3.0: JSON fields
        "Metric Formulas (JSON)": "metric_formulas",
        "Custom Properties (JSON)": "custom_properties",
    }

    # Add custom dimension fields
    for i in range(1, 11):
        field_mapping[f"Dim Custom {i}"] = f"dim_custom{i}"
        field_mapping[f"Dim Custom{i}"] = f"dim_custom{i}"  # Alternative format

    # Add custom cost fields
    for i in range(1, 11):
        field_mapping[f"Cost Custom {i}"] = f"cost_custom{i}"
        field_mapping[f"Cost Custom{i}"] = f"cost_custom{i}"  # Alternative format

    # Add custom metric fields
    for i in range(1, 11):
        field_mapping[f"Metric Custom {i}"] = f"metric_custom{i}"
        field_mapping[f"Metric Custom{i}"] = f"metric_custom{i}"  # Alternative format

    # NEW: Create mapping for calculated columns (for zero budget reconstruction)
    calculated_field_mapping = {}

    # Map percentage columns: "Cost Media %" -> "cost_media_pct"
    for header in headers:
        if " %" in header and not header.startswith("_"):  # Skip hidden preservation columns
            base_name = header.replace(" %", "")  # "Cost Media"

            # Map to corresponding cost field
            if base_name in ["Cost Media", "Cost Buying", "Cost Platform", "Cost Data", "Cost Creative"]:
                schema_field = f"cost_{base_name.lower().replace('cost ', '').replace(' ', '_')}_pct"
                calculated_field_mapping[header] = schema_field

            # Handle custom cost fields: "Cost Custom 1 %" -> "cost_custom1_pct"
            elif base_name.startswith("Cost Custom"):
                number = base_name.replace("Cost Custom ", "").replace("Cost Custom", "")
                schema_field = f"cost_custom{number}_pct"
                calculated_field_mapping[header] = schema_field

    # Map cost-per-unit columns: "Cost per Click" -> "metric_clicks_cpu"
    for header in headers:
        if header.startswith("Cost per ") and not header.startswith("_"):
            if "1000 Impressions" in header:
                calculated_field_mapping[header] = "metric_impressions_cpu"
            else:
                # Extract metric name from header
                metric_part = header.replace("Cost per ", "").lower()  # "click"

                # Map common metric names to schema fields (v3.0 comprehensive list)
                metric_mapping = {
                    "click": "metric_clicks_cpu",
                    "view": "metric_views_cpu",
                    "view start": "metric_view_starts_cpu",  # NEW v3.0
                    "view completion": "metric_view_completions_cpu",  # NEW v3.0
                    "reach": "metric_reach_cpu",  # NEW v3.0
                    "unit": "metric_units_cpu",  # NEW v3.0
                    "engagement": "metric_engagements_cpu",
                    "follower": "metric_followers_cpu",
                    "visit": "metric_visits_cpu",
                    "lead": "metric_leads_cpu",
                    "sale": "metric_sales_cpu",
                    "add to cart": "metric_add_to_cart_cpu",
                    "app install": "metric_app_install_cpu",
                    "application start": "metric_application_start_cpu",
                    "application complete": "metric_application_complete_cpu",
                    "contact us": "metric_contact_us_cpu",
                    "download": "metric_download_cpu",
                    "signup": "metric_signup_cpu",
                    "page view": "metric_page_views_cpu",  # NEW v3.0
                    "like": "metric_likes_cpu",  # NEW v3.0
                    "share": "metric_shares_cpu",  # NEW v3.0
                    "comment": "metric_comments_cpu",  # NEW v3.0
                    "conversion": "metric_conversions_cpu",  # NEW v3.0
                }

                if metric_part in metric_mapping:
                    calculated_field_mapping[header] = metric_mapping[metric_part]

                # Handle custom metrics: "Cost per Metric Custom 1" -> "metric_custom1_cpu"
                elif "metric custom" in metric_part:
                    number = metric_part.replace("metric custom ", "").replace("metric custom", "")
                    calculated_field_mapping[header] = f"metric_custom{number}_cpu"

    # Helper function to handle Excel errors and convert to appropriate values
    def clean_excel_value(cell_value, field_name: str):
        """Clean Excel cell values, converting errors to appropriate defaults."""
        if isinstance(cell_value, str) and cell_value.startswith('#'):
            logger.debug(f"Converting Excel error '{cell_value}' to 0 for field '{field_name}'")
            if field_name.startswith(("cost_", "metric_")) or field_name == "cost_total":
                return 0  # Numeric fields get 0
            else:
                return ""  # String fields get empty string

        if cell_value is None:
            return None

        return cell_value

    # Map column indices to field names
    header_to_field = {}
    calculated_header_to_field = {}
    skipped_columns = []

    for col_idx, header in enumerate(headers, 1):
        if header in field_mapping:
            header_to_field[col_idx] = field_mapping[header]
        elif header in calculated_field_mapping:
            calculated_header_to_field[col_idx] = calculated_field_mapping[header]
        elif any(pattern in str(header) for pattern in [" %", "Cost per "]):
            # This is a calculated column we don't recognize - skip it
            skipped_columns.append(header)

    # Log information about calculated vs skipped columns
    if calculated_header_to_field:
        logger.info(f"Importing {len(calculated_header_to_field)} calculated columns for zero budget reconstruction")
    if skipped_columns:
        logger.info(
            f"Skipping {len(skipped_columns)} unrecognized calculated columns: {', '.join(skipped_columns[:3])}")
        if len(skipped_columns) > 3:
            logger.info(f"... and {len(skipped_columns) - 3} more")

    # Process line items (skip header row)
    lineitems = []
    reconstructed_count = 0

    for row in range(2, line_items_sheet.max_row + 1):
        # Skip empty rows
        if all(cell.value is None for cell in line_items_sheet[row]):
            continue

        line_item = {}
        calculated_data = {}  # Store calculated column values temporarily

        # Process all original schema fields
        for col_idx, field_name in header_to_field.items():
            cell_value = line_items_sheet.cell(row=row, column=col_idx).value
            cleaned_value = clean_excel_value(cell_value, field_name)

            if cleaned_value is not None:
                # Handle different field types
                if field_name in ["id", "name"] and not cleaned_value:
                    if field_name == "id":
                        line_item[field_name] = f"li_{uuid.uuid4().hex[:8]}"
                    continue

                elif field_name in ["start_date", "end_date"]:
                    if isinstance(cleaned_value, date):
                        line_item[field_name] = cleaned_value.isoformat()
                    else:
                        line_item[field_name] = str(cleaned_value) if cleaned_value else ""

                # NEW v3.0: Handle JSON fields
                elif field_name in ["metric_formulas", "custom_properties"]:
                    line_item[field_name] = _parse_json_field(cleaned_value, f"lineitem.{field_name}")

                # NEW v3.0: Handle boolean fields
                elif field_name == "is_aggregate":
                    if isinstance(cleaned_value, bool):
                        line_item[field_name] = cleaned_value
                    elif isinstance(cleaned_value, str):
                        line_item[field_name] = cleaned_value.lower() in ["true", "yes", "1"]
                    else:
                        try:
                            line_item[field_name] = bool(cleaned_value)
                        except (ValueError, TypeError):
                            pass

                elif field_name.startswith(("cost_", "metric_")) or field_name == "cost_total":
                    try:
                        line_item[field_name] = float(cleaned_value)
                    except (ValueError, TypeError):
                        if field_name == "cost_total":
                            line_item[field_name] = 0  # cost_total is required

                elif field_name == "location_type":
                    if cleaned_value and str(cleaned_value).strip():
                        line_item[field_name] = str(cleaned_value).strip()

                else:
                    if str(cleaned_value).strip():
                        line_item[field_name] = str(cleaned_value).strip()

        # Process calculated fields (percentages and cost-per-unit) - only import for potential reconstruction
        for col_idx, calc_field_name in calculated_header_to_field.items():
            cell_value = line_items_sheet.cell(row=row, column=col_idx).value
            cleaned_value = clean_excel_value(cell_value, calc_field_name)

            if cleaned_value is not None:
                try:
                    calculated_data[calc_field_name] = float(cleaned_value)
                except (ValueError, TypeError):
                    pass

        # Ensure required fields have defaults
        if "id" not in line_item:
            line_item["id"] = f"li_{uuid.uuid4().hex[:8]}"
        if "name" not in line_item:
            line_item["name"] = line_item.get("id", "")
        if "start_date" not in line_item:
            line_item["start_date"] = "2025-01-01"
        if "end_date" not in line_item:
            line_item["end_date"] = "2025-12-31"
        if "cost_total" not in line_item:
            line_item["cost_total"] = 0

        # Apply zero budget reconstruction if needed
        cost_total = line_item.get('cost_total', 0)
        if cost_total == 0 and calculated_data:
            line_item = _reconstruct_zero_budget_breakdown(line_item, calculated_data)
            reconstructed_count += 1

        lineitems.append(line_item)

    # Log final results
    logger.info(f"Imported {len(lineitems)} line items")
    if reconstructed_count > 0:
        logger.info(
            f"Applied zero budget reconstruction to {reconstructed_count} line items using calculated column data")

    return lineitems


def _reconstruct_zero_budget_breakdown(line_item: Dict[str, Any], calculated_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Reconstruct cost breakdown and performance metrics from calculated column data when cost_total = 0.
    Assumes percentages are in decimal format (0.8 = 80%).

    Args:
        line_item: The line item with cost_total = 0
        calculated_data: Dictionary of calculated field values (percentages and cost-per-unit)

    Returns:
        Updated line item with reconstructed breakdown using minimal budget
    """
    name = line_item.get('name', line_item.get('id', 'Unknown'))
    minimum_budget = 0.0001

    # Check if we have meaningful calculated data to reconstruct from
    has_cost_percentages = any(key.endswith('_pct') and calculated_data[key] > 0 for key in calculated_data.keys())
    has_performance_cpu = any(key.endswith('_cpu') and calculated_data[key] > 0 for key in calculated_data.keys())

    if not (has_cost_percentages or has_performance_cpu):
        return line_item  # No meaningful calculated data to reconstruct from

    # Apply minimum budget
    line_item['cost_total'] = minimum_budget

    # Reconstruct cost breakdown from percentages (decimal format: 0.8 = 80%)
    if has_cost_percentages:
        total_percentage = 0
        cost_reconstructions = []

        for calc_field, percentage in calculated_data.items():
            if calc_field.endswith('_pct') and percentage > 0:
                # cost_media_pct -> cost_media
                base_field = calc_field.replace('_pct', '')

                # Percentage is already in decimal format (0.8 = 80%)
                cost_value = minimum_budget * percentage

                line_item[base_field] = cost_value
                cost_reconstructions.append(base_field)
                total_percentage += percentage

        logger.info(
            f"Reconstructed cost breakdown for '{name}': {len(cost_reconstructions)} cost components from {total_percentage * 100:.1f}% allocation")

    # Reconstruct performance metrics from cost-per-unit data
    if has_performance_cpu:
        metric_reconstructions = []

        for calc_field, cpu_value in calculated_data.items():
            if calc_field.endswith('_cpu') and cpu_value > 0:
                # metric_clicks_cpu -> metric_clicks
                base_field = calc_field.replace('_cpu', '')

                if base_field == 'metric_impressions':
                    # CPM calculation: impressions = (budget / cpu) * 1000
                    metric_value = (minimum_budget / cpu_value) * 1000
                else:
                    # Standard cost-per-unit: metric = budget / cpu
                    metric_value = minimum_budget / cpu_value

                line_item[base_field] = metric_value
                metric_reconstructions.append(base_field)

        logger.info(
            f"Reconstructed performance metrics for '{name}': {len(metric_reconstructions)} metrics from cost-per-unit ratios")

    # Log the reconstruction
    logger.info(f"Zero budget reconstruction for '{name}': ${minimum_budget:.4f} budget preserves all original ratios")

    return line_item


def _import_v3_dictionary(dictionary_sheet) -> Dict[str, Any]:
    """
    Import dictionary configuration for v3.0 (6 sections with formula support).

    Args:
        dictionary_sheet: The dictionary worksheet

    Returns:
        Dictionary containing v3.0 custom field and formula configuration
    """
    dictionary = {
        "meta_custom_dimensions": {},
        "campaign_custom_dimensions": {},
        "lineitem_custom_dimensions": {},
        "standard_metrics": {},
        "custom_metrics": {},
        "custom_costs": {}
    }

    # Find the headers (row 2)
    field_name_col = None
    field_type_col = None
    status_col = None
    caption_col = None
    formula_type_col = None
    base_metric_col = None

    # Look for header row (typically row 2)
    for row in range(1, 5):
        for col in range(1, 10):
            cell_value = dictionary_sheet.cell(row=row, column=col).value
            if cell_value == "Field Name":
                field_name_col = col
            elif cell_value == "Field Type":
                field_type_col = col
            elif cell_value == "Status":
                status_col = col
            elif cell_value == "Caption":
                caption_col = col
            elif cell_value == "Formula Type":
                formula_type_col = col
            elif cell_value == "Base Metric":
                base_metric_col = col

        # If we found the required headers, this is the header row
        if all(col is not None for col in [field_name_col, field_type_col]):
            break

    if not all(col is not None for col in [field_name_col, field_type_col]):
        logger.warning("Could not find required dictionary headers (Field Name, Field Type), returning empty dictionary")
        return {}

    # Process data rows (start from row 3)
    for row in range(3, dictionary_sheet.max_row + 1):
        field_name = dictionary_sheet.cell(row=row, column=field_name_col).value
        field_type = dictionary_sheet.cell(row=row, column=field_type_col).value

        if not field_name or not field_type:
            continue  # Skip incomplete rows

        # Normalize values
        field_name = str(field_name).strip()
        field_type = str(field_type).strip()

        # Get status and caption (for non-standard-metrics)
        status = None
        caption = None
        if status_col is not None:
            status_value = dictionary_sheet.cell(row=row, column=status_col).value
            if status_value:
                status = str(status_value).strip().lower()
        if caption_col is not None:
            caption_value = dictionary_sheet.cell(row=row, column=caption_col).value
            if caption_value:
                caption = str(caption_value).strip()

        # Get formula fields (for metrics)
        formula_type = None
        base_metric = None
        if formula_type_col is not None:
            formula_value = dictionary_sheet.cell(row=row, column=formula_type_col).value
            if formula_value and str(formula_value).strip():
                formula_type = str(formula_value).strip()
        if base_metric_col is not None:
            base_value = dictionary_sheet.cell(row=row, column=base_metric_col).value
            if base_value and str(base_value).strip():
                base_metric = str(base_value).strip()

        # Route to appropriate section based on field_type
        if field_type == "Meta Dimension":
            # Meta custom dimensions: dim_custom1-5
            if status and status in ["enabled", "disabled"]:
                config = {"status": status, "caption": caption if caption else ""}
                dictionary["meta_custom_dimensions"][field_name] = config

        elif field_type == "Campaign Dimension":
            # Campaign custom dimensions: dim_custom1-5
            if status and status in ["enabled", "disabled"]:
                config = {"status": status, "caption": caption if caption else ""}
                dictionary["campaign_custom_dimensions"][field_name] = config

        elif field_type == "LineItem Dimension":
            # LineItem custom dimensions: dim_custom1-10
            if status and status in ["enabled", "disabled"]:
                config = {"status": status, "caption": caption if caption else ""}
                dictionary["lineitem_custom_dimensions"][field_name] = config

        elif field_type == "Standard Metric":
            # Standard metrics: only formula_type and base_metric (NO status/caption)
            config = {}
            if formula_type:
                config["formula_type"] = formula_type
            if base_metric:
                config["base_metric"] = base_metric
            if config:  # Only add if at least one formula field is present
                dictionary["standard_metrics"][field_name] = config

        elif field_type == "Custom Metric":
            # Custom metrics: status, caption, formula_type, base_metric
            if status and status in ["enabled", "disabled"]:
                config = {"status": status, "caption": caption if caption else ""}
                if formula_type:
                    config["formula_type"] = formula_type
                if base_metric:
                    config["base_metric"] = base_metric
                dictionary["custom_metrics"][field_name] = config

        elif field_type == "Custom Cost":
            # Custom costs: status, caption
            if status and status in ["enabled", "disabled"]:
                config = {"status": status, "caption": caption if caption else ""}
                dictionary["custom_costs"][field_name] = config

    # Remove empty sections
    dictionary = {k: v for k, v in dictionary.items() if v}

    return dictionary


def _ensure_v3_meta_fields(meta: Dict[str, Any]) -> None:
    """
    Ensure required v3.0 meta fields are present with defaults.

    Args:
        meta: The meta dictionary to update
    """
    # v3.0: schema_version is required
    if "schema_version" not in meta:
        meta["schema_version"] = "3.0"

    # v3.0: id is required
    if "id" not in meta:
        meta["id"] = f"mediaplan_{uuid.uuid4().hex[:8]}"

    # v3.0: created_by_name is required
    if "created_by_name" not in meta:
        meta["created_by_name"] = meta.get("created_by", "Excel Import User")

    # v3.0: created_at is required
    if "created_at" not in meta:
        meta["created_at"] = datetime.now().isoformat()


def _ensure_v3_campaign_fields(campaign: Dict[str, Any]) -> None:
    """
    Ensure required v3.0 campaign fields are present with defaults.

    Args:
        campaign: The campaign dictionary to update
    """
    # Required campaign fields in v3.0
    if "id" not in campaign:
        campaign["id"] = f"campaign_{uuid.uuid4().hex[:8]}"
    if "name" not in campaign:
        campaign["name"] = "Campaign from Excel"
    if "objective" not in campaign:
        campaign["objective"] = "Imported from Excel"
    if "start_date" not in campaign:
        campaign["start_date"] = "2025-01-01"
    if "end_date" not in campaign:
        campaign["end_date"] = "2025-12-31"
    if "budget_total" not in campaign:
        campaign["budget_total"] = 100000

    # v3.0: Ensure target_audiences and target_locations arrays exist
    if "target_audiences" not in campaign:
        campaign["target_audiences"] = []
    if "target_locations" not in campaign:
        campaign["target_locations"] = []


def _sanitize_v3_media_plan_data(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Sanitize media plan data to avoid common validation errors for v3.0.

    Args:
        data: The media plan data to sanitize

    Returns:
        Sanitized media plan data
    """
    # Deep copy to avoid modifying the original
    sanitized = copy.deepcopy(data)

    # Handle campaign fields (v3.0: removed audience_gender and location_type from campaign)
    if "campaign" in sanitized:
        campaign = sanitized["campaign"]

        # Validate JSON fields
        if "custom_properties" in campaign:
            if not isinstance(campaign["custom_properties"], dict):
                campaign["custom_properties"] = {}

    # Handle lineitems
    if "lineitems" in sanitized:
        for line_item in sanitized["lineitems"]:
            # Handle location_type field (legacy compatibility - still in line items)
            if "location_type" in line_item and (line_item["location_type"] == "" or line_item["location_type"] is None):
                line_item["location_type"] = "Country"  # Default to Country

            # Validate JSON fields in line items
            if "metric_formulas" in line_item:
                if not isinstance(line_item["metric_formulas"], dict):
                    line_item["metric_formulas"] = {}
            if "custom_properties" in line_item:
                if not isinstance(line_item["custom_properties"], dict):
                    line_item["custom_properties"] = {}

    # Handle dictionary (v3.0: 6 sections with formula support)
    if "dictionary" in sanitized:
        dictionary = sanitized["dictionary"]

        # Validate and clean dictionary configuration for dimension sections
        for section_name in ["meta_custom_dimensions", "campaign_custom_dimensions", "lineitem_custom_dimensions", "custom_costs"]:
            if section_name in dictionary:
                section = dictionary[section_name]
                cleaned_section = {}

                for field_name, config in section.items():
                    if isinstance(config, dict) and "status" in config:
                        # Ensure status is valid
                        if config["status"] in ["enabled", "disabled"]:
                            cleaned_config = {
                                "status": config["status"],
                                "caption": config.get("caption", "")
                            }
                            # Ensure caption is provided for enabled fields
                            if config["status"] == "enabled" and not cleaned_config["caption"]:
                                cleaned_config["caption"] = f"Custom {field_name}"

                            cleaned_section[field_name] = cleaned_config

                if cleaned_section:
                    dictionary[section_name] = cleaned_section
                else:
                    # Remove empty sections
                    del dictionary[section_name]

        # Validate and clean standard_metrics section (formula_type and base_metric only)
        if "standard_metrics" in dictionary:
            section = dictionary["standard_metrics"]
            cleaned_section = {}

            for field_name, config in section.items():
                if isinstance(config, dict):
                    cleaned_config = {}
                    if "formula_type" in config and config["formula_type"]:
                        cleaned_config["formula_type"] = config["formula_type"]
                    if "base_metric" in config and config["base_metric"]:
                        cleaned_config["base_metric"] = config["base_metric"]

                    if cleaned_config:  # Only add if at least one formula field
                        cleaned_section[field_name] = cleaned_config

            if cleaned_section:
                dictionary["standard_metrics"] = cleaned_section
            else:
                del dictionary["standard_metrics"]

        # Validate and clean custom_metrics section (status, caption, formula_type, base_metric)
        if "custom_metrics" in dictionary:
            section = dictionary["custom_metrics"]
            cleaned_section = {}

            for field_name, config in section.items():
                if isinstance(config, dict) and "status" in config:
                    if config["status"] in ["enabled", "disabled"]:
                        cleaned_config = {
                            "status": config["status"],
                            "caption": config.get("caption", "")
                        }

                        # Ensure caption is provided for enabled fields
                        if config["status"] == "enabled" and not cleaned_config["caption"]:
                            cleaned_config["caption"] = f"Custom {field_name}"

                        # Add formula fields if present
                        if "formula_type" in config and config["formula_type"]:
                            cleaned_config["formula_type"] = config["formula_type"]
                        if "base_metric" in config and config["base_metric"]:
                            cleaned_config["base_metric"] = config["base_metric"]

                        cleaned_section[field_name] = cleaned_config

            if cleaned_section:
                dictionary["custom_metrics"] = cleaned_section
            else:
                del dictionary["custom_metrics"]

    return sanitized